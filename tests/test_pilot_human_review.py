"""
Tests for llm-sum/pilot_human_review.py — the pilot human-validation harness.

IN PLAIN ENGLISH: this file doesn't run the real program on real data. Instead,
each test builds a small fake, throwaway version of the project's data files
(fake articles, fake AI summaries, fake scores) inside a temporary folder,
runs pilot_human_review.py's export function against that fake data, and then
checks that the files it produced look exactly as expected. Because nothing
here calls a real AI model or the internet, it's completely safe to run.

Offline, mock corpus only (PHASE3_MODE=test, $0, no network). Covers:
    - The relocated dev-pool DOI reader (eval_instances.read_dois_from_dev_folder).
    - A self-contained humanN/ folder: guide + packet + .xlsx + one article.pdf
      per item folder (copied from the source PDF).
    - The blind protocol: packet.md AND the .xlsx carry no summariser identity.
    - packet.md points at the actual .xlsx scoresheet, not the CSV default.
    - Partial overlap: human2 shares exactly round(N*ratio) of human1's items.
    - human2's export never mutates human1's folder.
    - Prerequisite handling: summarized-but-unevaluated warning; empty pool -> None.
    - Per-item PDF copy, including the graceful fallback to article.md + warning
      when a PDF is absent.
    - Pool-exhaustion backfill warning when the dev pool can't supply new items.
"""

# Makes newer type-hint syntax (e.g. `int | None`) work on older Python
# versions; has no effect on the tests themselves.
from __future__ import annotations

# The lines below are "import" statements: each one loads a ready-made
# toolbox of code (either Python's standard library, a third-party package,
# or another file in this project) so this test file can reuse it.
import json  # reads/writes .json files (the private "unblinding key")
from pathlib import Path  # represents file/folder paths

import pytest  # the testing framework that discovers and runs test_ functions

import eval_instances
import human_review
import prepare_texts
from evaluator import BLIND_FORBIDDEN_TOKENS
from file_paths import descriptive_pdf_filename, doi_to_slug
import pilot_human_review
from pilot_human_review import export_pilot_human_review

from openpyxl import load_workbook  # reads .xlsx spreadsheet files, to check the scoresheet


# Sample journal names and AI-provider names, reused across the fake test
# data built by every test below (so `providers=PROVIDERS` etc. stay
# consistent without repeating the literal lists everywhere). JOURNALS MUST
# be the real study journals -- the pilot's sampler always draws evenly
# across human_review.STUDY_JOURNALS internally (dev-pool scoping is its
# only difference from the real export), so a fixture using different
# journal names would leave every real journal short and trigger backfill
# on every draw.
JOURNALS = human_review.STUDY_JOURNALS
PROVIDERS = ("openai", "anthropic", "gemini")


# In plain English: pretends a "summarize --mode dev" run already happened
# for one article, by writing the small marker .txt file that step leaves
# behind. This is what makes an article "eligible" for the pilot pool in the
# tests below — pilot_human_review only samples from what it can find here.
def _write_dev_txt(folder: Path, doi: str, *, suffix: str = "run") -> None:
    folder.mkdir(parents=True, exist_ok=True)
    (folder / f"{doi_to_slug(doi)}__{suffix}.txt").write_text(
        f"DOI: {doi}\nJournal: x\n\nbody\n", encoding="utf-8",
    )


# This is the main "build a fake mini-project" helper that almost every test
# below starts with. ``tmp_path`` and ``monkeypatch`` are pytest "fixtures" —
# ready-made helpers pytest automatically hands to any test/helper function
# that asks for them by name: ``tmp_path`` is a brand-new, empty temporary
# folder that gets thrown away after the test, and ``monkeypatch`` lets a
# test temporarily redirect a piece of code (here, a file-path setting) to
# point somewhere else, automatically undoing the change once the test ends.
# Using these means the tests never touch the project's real data/ folder.
def _build_fixture(
    tmp_path: Path, monkeypatch, *,
    num_articles: int = 5,
    providers: tuple[str, ...] = PROVIDERS,
    with_pdfs: bool = True,
    missing_pdf_index: int | None = None,
    dev_dois: "set[str] | None" = None,
) -> dict:
    """Build an isolated dev-scoped corpus and return the paths to drive the export.

    In plain English: creates a small pretend version of everything
    pilot_human_review.py normally reads from the real project — fake
    articles, fake AI-written summaries, fake LLM-judge scores, fake source
    PDFs, and a fake "these articles were dev-summarized" marker folder —
    all inside the disposable ``tmp_path`` folder, then hands back a
    dictionary of paths pointing at all of it so a test can call the export
    function against this fake data instead of anything real.

    ``num_articles`` distinct DOIs (one per journal, cycling), each summarized +
    evaluated by every ``providers`` entry. Reference text is cached so the
    text-join succeeds. ``with_pdfs`` drops a dummy PDF in raw/ for each article
    (skip ``missing_pdf_index`` to exercise the missing-PDF path). ``dev_dois``
    overrides which DOIs land in the dev-summary folder (defaults to all).
    """
    # --- Set up fake versions of the folders the real pipeline would use ---
    processed_dir = tmp_path / "processed"
    processed_dir.mkdir()
    monkeypatch.setattr(prepare_texts, "PROCESSED_DIR", processed_dir)
    # Keep the manifest join hermetic — no real data/manifest.jsonl.
    monkeypatch.setattr(eval_instances, "MANIFEST_PATH", tmp_path / "no_manifest.jsonl")
    monkeypatch.setattr(eval_instances, "MANUAL_MANIFEST_PATH", tmp_path / "no_manual.jsonl")

    summaries_path = tmp_path / "summaries.jsonl"
    evaluations_path = tmp_path / "evaluations.jsonl"
    raw_dir = tmp_path / "raw"
    raw_dir.mkdir()
    dev_dir = tmp_path / "dev_summaries_jsonl"

    # --- Make up ``num_articles`` fake articles, one journal each (cycling) ---
    dois = [f"10.1111/test.multi{i}" for i in range(num_articles)]
    records = []
    for i, doi in enumerate(dois):
        journal = JOURNALS[i % len(JOURNALS)]
        title = f"Study {i} of Condition in Species {i}"
        record = {"doi": doi, "journal": journal, "title": title}
        records.append(record)

    # --- Write a fake summaries.jsonl: one row per article, one AI summary
    # per provider (openai/anthropic/gemini) ---
    # summaries.jsonl (one row per article, every provider a success slot).
    # "with open(path, "w", ...) as f:" opens a file for writing and
    # automatically closes it again once the indented block below finishes —
    # even if something goes wrong partway through.
    with open(summaries_path, "w", encoding="utf-8") as f:
        for record in records:
            row = dict(record)
            # Summary text carries NO provider name — a real summary wouldn't,
            # and the blind check asserts reviewer-facing text is model-agnostic.
            row["models"] = {
                p: {"status": "success",
                    "summary": f"Candidate summary variant {j} of {record['doi']}."}
                for j, p in enumerate(providers)
            }
            f.write(json.dumps(row) + "\n")

    # --- Write a fake "cleaned article text" file for each article, so the
    # code that joins a score row back to its article text can find it ---
    # Cached reference text so the text-join resolves.
    for record in records:
        slug = doi_to_slug(record["doi"])
        (processed_dir / f"{slug}.jsonl").write_text(
            json.dumps({"doi": record["doi"], "slug": slug,
                        "text": f"Full cleaned article text for {record['doi']}."}) + "\n",
            encoding="utf-8",
        )

    # --- Write a fake evaluations.jsonl: one judged score row per
    # (article, AI provider) pair ---
    # evaluations.jsonl (one row per article x provider).
    with open(evaluations_path, "w", encoding="utf-8") as f:
        for i, record in enumerate(records):
            for provider in providers:
                f.write(json.dumps({
                    "doi": record["doi"],
                    "summarizer": provider,
                    "judge": "anthropic",
                    "input_source": "processed",
                    "rubric_version": "vet_medhelm_score_v1.0",
                    "jury_score": 4.0,
                    "jury_score_weighted": 4.0,
                    "jury_score_unweighted": 4.0,
                    "criteria_scores": {"faithfulness": {"score": 4, "reasoning": "ok"}},
                    "requires_human_review": False,
                    "strata": {"species": ["Canine"], "study_design": "RCT",
                               "clinical_topic": "Cardiology", "journal": record["journal"],
                               "input_source": "processed"},
                    "timestamp": "2026-01-01T00:00:00+00:00",
                }) + "\n")

    # --- Drop a dummy (fake, non-real) PDF for each article, so the "copy
    # source PDFs into the reviewer folder" step has something to find ---
    # Dummy source PDFs (descriptive name, matching resolve_existing_pdf_path).
    if with_pdfs:
        for i, record in enumerate(records):
            if missing_pdf_index is not None and i == missing_pdf_index:
                continue
            (raw_dir / descriptive_pdf_filename(record)).write_bytes(b"%PDF-1.4 dummy\n")

    # --- Mark which articles count as "already dev-summarized", i.e.
    # eligible for the pilot pool ---
    # Dev-summary folder scopes the eligible pool.
    scoped = dev_dois if dev_dois is not None else set(dois)
    for doi in scoped:
        _write_dev_txt(dev_dir, doi)

    # Redirect human_review.py's own path constants at this fake data too, so
    # the shared text-joining code it exposes reads the fake files, not the
    # real project's.
    # Point the shared text-source constants at the fixture.
    monkeypatch.setattr(human_review, "SUMMARIES_PATH", summaries_path)
    monkeypatch.setattr(human_review, "SUMMARIES_TXT_DIR", tmp_path / "no_summaries_txt")
    monkeypatch.setattr(human_review, "DEV_TESTS_SUMMARIES_TXT_DIR", tmp_path / "no_dev_txt")

    # Hand back every path a test might need, plus the raw article records,
    # bundled in a dict (a labeled lookup table) for convenience.
    return {
        "summaries_path": summaries_path,
        "evaluations_path": evaluations_path,
        "raw_dir": raw_dir,
        "dev_dir": dev_dir,
        "output_dir": tmp_path / "pilot_human_review",
        "dois": dois,
        "records": records,
    }


# A small convenience wrapper: takes the fixture dict from _build_fixture()
# and calls the real export_pilot_human_review() function, pre-filling all
# the fake-data paths so each test only needs to specify what it actually
# cares about varying (e.g. `overlap_ratio=0.6`). ``**overrides`` means "any
# extra named arguments the caller passes get merged in", overriding the
# defaults if there's a name clash.
def _export(fx: dict, **overrides):
    kwargs = dict(
        evaluations_path=fx["evaluations_path"],
        dev_summaries_dir=fx["dev_dir"],
        raw_dir=fx["raw_dir"],
        output_dir=fx["output_dir"],
    )
    kwargs.update(overrides)
    return export_pilot_human_review(**kwargs)


# Reads a written unblinding-key .json file and returns the set of
# (doi, summarizer, input_source) identities it contains — i.e. "which
# articles, written by which AI, does this tester's folder actually cover?"
# Used by several tests to compare what human1 vs human2 were shown.
def _key_identities(key_path: Path) -> set[tuple[str, str, str]]:
    key = json.loads(key_path.read_text(encoding="utf-8"))
    return {
        (v["doi"], v["summarizer"], v.get("input_source", "processed"))
        for v in key["items"].values()
    }


# ---------------------------------------------------------------------------
# Relocated dev-pool DOI reader
# ---------------------------------------------------------------------------

# In plain English: checks the little function that scans the
# dev_summaries_jsonl/ folder and reports back which articles (by DOI) have
# a marker file in it. Two files for the same article should only count once,
# and a folder that doesn't exist should report "no articles", not crash.
def test_read_dois_from_dev_folder_parses_headers(tmp_path: Path) -> None:
    folder = tmp_path / "dev_summaries_jsonl"
    _write_dev_txt(folder, "10.1/aaa", suffix="run_A")
    _write_dev_txt(folder, "10.1/aaa", suffix="run_B")  # same DOI, new suffix
    _write_dev_txt(folder, "10.2/bbb")
    assert eval_instances.read_dois_from_dev_folder(folder) == {"10.1/aaa", "10.2/bbb"}
    assert eval_instances.read_dois_from_dev_folder(tmp_path / "missing") == set()


# ---------------------------------------------------------------------------
# Self-contained folder + blind protocol
# ---------------------------------------------------------------------------

# In plain English: runs a full pilot export once and checks that the very
# first tester's folder ("human1") contains everything a reviewer would need
# — the guide, the packet index, the spreadsheet, one subfolder per scored
# item — and that the private answer key ends up NEXT TO the human1 folder,
# never inside it (so handing the folder to a real person can't leak
# anything). Also checks the expected 15 items (5 articles x 3 AI providers)
# and that every item folder got its own copy of the source PDF.
def test_pilot_export_creates_self_contained_human1(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    result = _export(fx, overlap_ratio=0.6, seed=42)

    assert result is not None
    assert result.human_number == 1
    # Default unit is now "articles": 5 evaluated dev articles x 3 providers.
    assert result.items_exported == 15
    assert result.overlap_units == 0

    human_dir = fx["output_dir"] / "human1"
    assert (human_dir / "REVIEWER_GUIDE.md").exists()
    assert (human_dir / "packet.md").exists()
    assert (human_dir / "scoresheet_human1.xlsx").exists()
    # Private key is a SIBLING of the folder, never inside it.
    assert (fx["output_dir"] / "unblinding_key_human1.json").exists()
    assert not (human_dir / "unblinding_key_human1.json").exists()

    # Nested per-item layout: one item_NNN/ folder per item, each with its own
    # article.pdf (the source PDF, copied per item) + summary.md. The
    # top-level original_articles/ dedup folder is gone now that every item
    # carries its own article.pdf.
    item_folders = sorted(p for p in human_dir.glob("item_*") if p.is_dir())
    assert len(item_folders) == 15
    for folder in item_folders:
        assert (folder / "article.pdf").exists()
        assert not (folder / "article.md").exists()
        assert (folder / "summary.md").exists()
    assert not (human_dir / "original_articles").exists()


# In plain English: gathers up every piece of text a reviewer would actually
# read in a humanN/ folder (the packet, the guide, every article.pdf/.md and
# summary.md) into one big lowercase blob, so a test can search it for
# forbidden words in one go. article.pdf is binary, so it's decoded loosely
# (ignoring bytes that aren't plain ASCII) rather than as UTF-8 text.
def _folder_text(human_dir: Path) -> str:
    """All reviewer-facing text in a humanN/ folder (packet + guide + item files)."""
    parts = [
        (human_dir / "packet.md").read_text(encoding="utf-8"),
        (human_dir / "REVIEWER_GUIDE.md").read_text(encoding="utf-8"),
    ]
    for folder in human_dir.glob("item_*"):
        article_pdf = folder / "article.pdf"
        if article_pdf.exists():
            parts.append(article_pdf.read_bytes().decode("ascii", errors="ignore"))
        else:
            parts.append((folder / "article.md").read_text(encoding="utf-8"))
        parts.append((folder / "summary.md").read_text(encoding="utf-8"))
    return "\n".join(parts).lower()


# In plain English: checks two things at once — (1) the packet.md file
# correctly points readers at the real .xlsx spreadsheet filename (not the
# CSV name the real, non-pilot export would use), and (2) the "blind
# protocol" holds: nowhere in any reviewer-facing text or in the spreadsheet
# itself does an AI provider's name leak through (see BLIND_FORBIDDEN_TOKENS
# in evaluator.py for the list of forbidden words).
def test_pilot_packet_references_xlsx_and_is_blind(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    _export(fx, overlap_ratio=0.6, seed=42)
    human_dir = fx["output_dir"] / "human1"

    packet = (human_dir / "packet.md").read_text(encoding="utf-8")
    assert "scoresheet_human1.xlsx" in packet
    assert "scoresheet_reviewer" not in packet  # not the CSV default
    # The packet is now a navigation index pointing at per-item folders.
    assert "item_001/" in packet
    assert "Study 1 of Condition" in packet  # blind-safe article title listed
    # The full article text + candidate summary live in the item folders now.
    summary_md = (human_dir / "item_001" / "summary.md").read_text(encoding="utf-8")
    assert "**Article:**" in summary_md

    all_text = _folder_text(human_dir)
    wb = load_workbook(human_dir / "scoresheet_human1.xlsx")
    ws = wb.active
    xlsx_text = " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    ).lower()
    for forbidden in BLIND_FORBIDDEN_TOKENS:
        assert forbidden not in all_text
        assert forbidden not in xlsx_text


# In plain English: reads the private key file and returns just the list of
# article DOIs, in the same order the items appear (item_001, item_002, ...).
def _ordered_dois(key_path: Path) -> list[str]:
    """DOIs in packet/scoresheet order (item_001, item_002, ...)."""
    key = json.loads(key_path.read_text(encoding="utf-8"))
    items = key["items"]
    return [items[k]["doi"] for k in sorted(items)]


# In plain English: checks the ORDER items get laid out in. With 5 articles
# and 3 AI providers, item 1-5 should be five different articles (each
# paired with the "first" provider); items 6-10 the same five articles again
# paired with the second provider; items 11-15 the third. It also checks
# that the exact same article is never scored twice back-to-back, since a
# reviewer scoring the same article twice in a row could unconsciously let
# the first score bias the second.
def test_pilot_items_cycle_provider_major(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    _export(fx, seed=42, sample_size=5)
    dois = _ordered_dois(fx["output_dir"] / "unblinding_key_human1.json")
    assert len(dois) == 15

    # Round 0 (items 1-5) is five distinct articles; each recurs 5 and 10 later
    # (same within-round slot) as the next provider's summary -> provider-major
    # cycling: article1..5 provider A, then article1..5 provider B, then C.
    assert len(set(dois[:5])) == 5
    for i in range(5):
        assert dois[i] == dois[i + 5] == dois[i + 10]
    # And no two adjacent items are the same article (independent scoring).
    for i in range(len(dois) - 1):
        assert dois[i] != dois[i + 1]


# In plain English: since the same article gets scored three times (once per
# AI provider), the spreadsheet labels each repeat "summary version 1 of 3",
# "version 2 of 3", etc. so a reviewer can tell the rows apart without being
# told which AI wrote which. This test checks every item_id is unique and
# that the version labels come out exactly 5-5-5 across the 15 rows.
def test_pilot_scoresheet_labels_summary_versions(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    _export(fx, seed=42, sample_size=5)
    wb = load_workbook(fx["output_dir"] / "human1" / "scoresheet_human1.xlsx")
    ws = wb.active
    title_col = human_review.SCORESHEET_FIELDS.index("article_title") + 1
    id_col = human_review.SCORESHEET_FIELDS.index("item_id") + 1
    titles = [ws.cell(row=r, column=title_col).value for r in range(2, 2 + 15)]
    ids = [ws.cell(row=r, column=id_col).value for r in range(2, 2 + 15)]

    assert len(set(ids)) == 15  # every row keyed by a distinct item_id
    # Each of the 5 articles contributes three rows, labeled version 1/2/3 of 3.
    for k in (1, 2, 3):
        assert sum(1 for t in titles if f"summary version {k} of 3" in t) == 5


# In plain English: checks the spreadsheet's usability touches — the header
# row stays visible while scrolling ("frozen panes"), the column headers
# match the expected field names, and the two dropdown pickers exist (1-5
# for scores, yes/no for the hallucination question), so a reviewer can't
# accidentally type an invalid value.
def test_pilot_xlsx_has_dropdowns_and_frozen_header(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    _export(fx, seed=42)
    wb = load_workbook(fx["output_dir"] / "human1" / "scoresheet_human1.xlsx")
    ws = wb.active
    assert ws.freeze_panes == "A2"
    header = [ws.cell(row=1, column=i + 1).value for i in range(len(human_review.SCORESHEET_FIELDS))]
    assert header == human_review.SCORESHEET_FIELDS
    # Two data validations: 1-5 for the criteria columns, yes/no for hallucination.
    formulas = {dv.formula1 for dv in ws.data_validations.dataValidation}
    assert '"1,2,3,4,5"' in formulas
    assert '"yes,no"' in formulas


# ---------------------------------------------------------------------------
# Incremental folders + partial overlap
# ---------------------------------------------------------------------------

# In plain English: like _key_identities, but returns just the article DOIs
# (dropping which AI wrote each summary) — useful when a test only cares
# about which ARTICLES overlapped between two testers, not every summary.
def _key_dois(key_path: Path) -> set[str]:
    return {doi for doi, _, _ in _key_identities(key_path)}


# In plain English: the core overlap-ratio test. Exports human1, then human2
# with overlap_ratio=0.6 (60%), and checks the math works out exactly:
# with 5 sampled articles, round(5 x 0.6) = 3 should be repeated for human2
# and 2 should be brand-new, which (since each article carries 3 AI
# summaries) means 9 of human2's 15 items overlap with human1's and 6 don't.
def test_pilot_overlap_shares_expected_fraction(tmp_path: Path, monkeypatch) -> None:
    # 8 articles/journal (40 total) so human2 can carry 3/journal and still
    # find 2 genuinely-new articles/journal without exhausting the pool.
    # 1 provider keeps "articles" and "items" counts identical and readable.
    fx = _build_fixture(tmp_path, monkeypatch, num_articles=8 * len(JOURNALS), providers=("openai",))
    r1 = _export(fx, overlap_ratio=0.6, seed=42, sample_size=25)
    r2 = _export(fx, overlap_ratio=0.6, seed=42, sample_size=25)

    assert r1.human_number == 1 and r2.human_number == 2
    i1 = _key_identities(fx["output_dir"] / "unblinding_key_human1.json")
    i2 = _key_identities(fx["output_dir"] / "unblinding_key_human2.json")
    # 25 articles x 1 provider each.
    assert len(i1) == 25 and len(i2) == 25
    # Overlap is carried per journal: round(5/journal * 0.6) == 3 shared
    # articles/journal (=15 shared across 5 journals), 2 new/journal (=10 new).
    d1 = _key_dois(fx["output_dir"] / "unblinding_key_human1.json")
    d2 = _key_dois(fx["output_dir"] / "unblinding_key_human2.json")
    assert len(d1 & d2) == 15
    assert len(d2 - d1) == 10
    assert len(i1 & i2) == 15
    assert len(i2 - i1) == 10
    assert r2.overlap_units == 15


# In plain English: the "1.0" regime — every new tester should see EXACTLY
# the same items as the previous one, with nothing new mixed in.
def test_pilot_overlap_ratio_one_reuses_all(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    _export(fx, overlap_ratio=1.0, seed=42)
    _export(fx, overlap_ratio=1.0, seed=42)
    i1 = _key_identities(fx["output_dir"] / "unblinding_key_human1.json")
    i2 = _key_identities(fx["output_dir"] / "unblinding_key_human2.json")
    assert i1 == i2  # identical items every run


# In plain English: the "0.0" regime — every new tester should get a
# completely fresh, non-overlapping set of items, sharing nothing with the
# previous tester ("isdisjoint" means "the two sets have no items in common").
def test_pilot_overlap_ratio_zero_is_fresh_draw(tmp_path: Path, monkeypatch) -> None:
    # 10 articles in the pool so a ratio-0 second draw of 5 can be fully disjoint
    # (with only 5 articles, human2 would exhaust the pool and backfill).
    fx = _build_fixture(tmp_path, monkeypatch, num_articles=10)
    _export(fx, overlap_ratio=0.0, seed=42, sample_size=5)
    _export(fx, overlap_ratio=0.0, seed=42, sample_size=5)
    i1 = _key_identities(fx["output_dir"] / "unblinding_key_human1.json")
    i2 = _key_identities(fx["output_dir"] / "unblinding_key_human2.json")
    assert i1.isdisjoint(i2)  # no carried items


# In plain English: safety check that creating human2's folder never
# rewrites anything inside human1's — it records human1's packet.md content
# and its file-modified timestamp beforehand, creates human2, then confirms
# human1's file is byte-for-byte and timestamp-for-timestamp unchanged.
def test_pilot_human2_does_not_touch_human1(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    _export(fx, overlap_ratio=0.6, seed=42)
    human1_packet = fx["output_dir"] / "human1" / "packet.md"
    before = human1_packet.read_text(encoding="utf-8")
    before_mtime = human1_packet.stat().st_mtime_ns

    _export(fx, overlap_ratio=0.6, seed=42)  # creates human2

    assert human1_packet.read_text(encoding="utf-8") == before
    assert human1_packet.stat().st_mtime_ns == before_mtime
    assert (fx["output_dir"] / "human2" / "packet.md").exists()


# ---------------------------------------------------------------------------
# Prerequisites + empty pool
# ---------------------------------------------------------------------------
# In plain English: this section checks the export behaves sensibly when the
# data isn't fully ready yet — e.g. an article was summarized but not yet
# judged, or there's nothing eligible at all.

# In plain English: adds one extra "dev-summarized" article that has NO
# evaluation score yet, and checks the export (a) still succeeds for the
# other 5 properly-evaluated articles, and (b) prints a clear warning naming
# the unready article instead of silently ignoring or crashing on it.
# ``capsys`` is a pytest fixture that captures whatever the code being tested
# printed to the terminal, so the test can inspect that text.
def test_pilot_warns_when_dev_doi_has_no_eval_rows(tmp_path: Path, monkeypatch, capsys) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    # Add a dev summary for a DOI that has no evaluation rows.
    _write_dev_txt(fx["dev_dir"], "10.1111/test.unevaluated")
    result = _export(fx, seed=42)
    out = capsys.readouterr().out
    assert "summarized but have no rows" in out
    assert "10.1111/test.unevaluated" in out
    # The evaluated articles still export fine (5 articles x 3 providers).
    assert result is not None and result.items_exported == 15


# In plain English: if the ONLY dev article available has no evaluation
# score, there's nothing samplable — the export should cleanly return
# "nothing to do" (None) and must not create a half-finished human1/ folder.
def test_pilot_returns_none_when_no_evaluated_dev_articles(tmp_path: Path, monkeypatch) -> None:
    # Dev pool points only at a DOI with no eval rows -> empty eligible pool.
    fx = _build_fixture(tmp_path, monkeypatch, dev_dois={"10.1111/test.orphan"})
    result = _export(fx, seed=42)
    assert result is None
    assert not (fx["output_dir"] / "human1").exists()


# In plain English: if "summarize --mode dev" has never been run at all (the
# dev_summaries_jsonl/ folder is completely empty), the export should return
# "nothing to do" rather than error out.
def test_pilot_returns_none_when_dev_folder_empty(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch, dev_dois=set())
    result = _export(fx, seed=42)
    assert result is None


# ---------------------------------------------------------------------------
# PDF matching
# ---------------------------------------------------------------------------

# In plain English: with overlap_ratio=1.0 (so human1 gets every fixture
# article), checks that a real PDF got copied into every item's own
# article.pdf, and nothing was reported missing.
def test_pilot_copies_matched_pdfs(tmp_path: Path, monkeypatch) -> None:
    fx = _build_fixture(tmp_path, monkeypatch)
    result = _export(fx, overlap_ratio=1.0, seed=42)
    human_dir = fx["output_dir"] / "human1"
    item_folders = [p for p in human_dir.glob("item_*") if p.is_dir()]
    copied = [p for p in item_folders if (p / "article.pdf").exists()]
    # Every item folder got its own copy of its article's source PDF.
    assert result.pdfs_copied == len(copied) == len(item_folders) == result.items_exported
    assert result.pdfs_missing == []
    for folder in item_folders:
        assert not (folder / "article.md").exists()


# In plain English: builds a fixture where one article's PDF is deliberately
# missing from raw/, and checks the export doesn't crash — it should still
# finish, list that article's item(s) under "pdfs_missing" (falling back to
# article.md for just those items), print a warning mentioning it, and still
# successfully copy the other articles' PDFs.
def test_pilot_missing_pdf_warns_not_fatal(tmp_path: Path, monkeypatch, capsys) -> None:
    # Force the whole 5-item sample so the article with no PDF is always included.
    fx = _build_fixture(tmp_path, monkeypatch, missing_pdf_index=0)
    result = _export(fx, seed=42, sample_size=5)
    out = capsys.readouterr().out
    assert result is not None
    assert result.pdfs_missing  # the missing article's item(s) are reported
    assert "no source PDF found" in out
    # The rest still copied.
    assert result.pdfs_copied >= 1

    human_dir = fx["output_dir"] / "human1"
    item_folders = [p for p in human_dir.glob("item_*") if p.is_dir()]
    pdf_items = [p for p in item_folders if (p / "article.pdf").exists()]
    md_items = [p for p in item_folders if (p / "article.md").exists()]
    # Every item falls back to exactly one of the two — never both, never neither.
    assert len(pdf_items) + len(md_items) == len(item_folders) == result.items_exported
    assert len(pdf_items) == result.pdfs_copied
    assert len(md_items) == len(result.pdfs_missing)


# ---------------------------------------------------------------------------
# Pool exhaustion
# ---------------------------------------------------------------------------

# In plain English: what happens when the dev pool is so small that there
# just aren't enough genuinely-new articles left to give the next tester?
# With exactly 1 article/journal (the smallest valid quota), human1 already
# used every article that exists; human2 asks for a fresh, non-overlapping
# draw (ratio 0.0) but there is no "new" article left in ANY journal. This
# checks the export doesn't under-fill or crash — it "backfills" by reusing
# an already-seen article and prints a clear warning explaining why, instead
# of silently shortchanging the tester.
def test_pilot_pool_exhaustion_backfills_with_warning(tmp_path: Path, monkeypatch, capsys) -> None:
    # Exactly 1 article/journal (5 total), 1 provider each -> the smallest
    # valid quota (sample_size=5) exactly exhausts the pool.
    fx = _build_fixture(tmp_path, monkeypatch, num_articles=len(JOURNALS), providers=("openai",))
    _export(fx, overlap_ratio=0.0, seed=42, sample_size=5)
    capsys.readouterr()  # drop human1 output
    r2 = _export(fx, overlap_ratio=0.0, seed=42, sample_size=5)
    out = capsys.readouterr().out
    assert r2 is not None
    assert r2.items_exported == 5  # not under-filled
    assert "could not supply" in out and "backfilled" in out
