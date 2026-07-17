"""
Tests for llm-sum/human_review.py — human validation sampling + blind export.

Critical assertions:
    Blind check: rendered packet + scoresheet contain NONE of the forbidden
    tokens (mirrors evaluator.py's blind judge prompt test).
    Sampler: rows flagged requires_human_review are prioritized; remaining
    slots are filled with a stratified, deterministic (seeded) sample.
    Export: each call writes one more humanN/ folder, exactly split across
    the study's 5 journals; the per-reviewer un-blinding key is the only
    file carrying identity.
    Rows whose source text can't be resolved are skipped, not fatal.

IN PLAIN ENGLISH
-----------------
This file is a set of automated checks ("tests") that make sure
llm-sum/human_review.py's EXPORT half (the part that builds reviewer
folders — see that file's own top-of-file explanation) keeps working
correctly. Each function below whose name starts with ``test_`` is one
independent check: it sets up a small, fake/example situation, runs a piece
of human_review.py against it, and then uses ``assert`` statements to verify
the result was what's expected. ``assert <condition>`` means "crash loudly
with an error if this is not true" — that's how a test reports failure.
These tests never call a real AI or the internet; everything runs against
small, hand-built fake data, which is why they're safe to run automatically.
"""

from __future__ import annotations

# Standard-library imports: csv/io/json for reading and building test data
# files in memory, Path for filesystem locations. See human_review.py's own
# imports section for a one-line explanation of each of these.
import csv
import io
import json
from pathlib import Path

# Imports from this project's own code — the pieces under test, plus a
# couple of helpers reused to build realistic fake data.
import prepare_texts
from evaluator import BLIND_FORBIDDEN_TOKENS
from file_paths import descriptive_pdf_filename, doi_to_slug
import human_review
from human_review import (
    REVIEWER_GUIDE_PATH,
    STUDY_JOURNALS,
    ReviewItem,
    SCORESHEET_FIELDS,
    build_unblinding_key,
    export_human_review,
    render_packet_markdown,
    render_reviewer_guide_markdown,
    render_scoresheet_csv,
    sample_rows_for_review,
)


# ---------------------------------------------------------------------------
# Blind protocol
# ---------------------------------------------------------------------------
# In plain English: the tests in this section check the single most
# important safety rule in the whole file — that nothing shown to a
# reviewer ever reveals which AI wrote a given summary.

def _make_item(item_id: str, *, summarizer: str = "openai", judge: str = "anthropic",
               title: str = "A Retrospective Study of Condition X in Dogs") -> ReviewItem:
    """Not a test itself — a helper used BY the tests below to quickly build
    one fully-filled-in ReviewItem (see human_review.py's ReviewItem
    dataclass) with sensible fake values, so each test doesn't have to
    repeat all ten-plus fields by hand. Any field can be overridden via a
    keyword argument (e.g. ``_make_item("item_002", summarizer="gemini")``).
    """
    return ReviewItem(
        item_id=item_id,
        doi="10.1111/test.review",
        title=title,
        summarizer=summarizer,
        judge=judge,
        rubric_version="vet_medhelm_score_v1.0",
        input_source="processed",
        reference_text="Dogs with condition X improved after treatment Y in this retrospective study.",
        candidate_summary="This study found that treatment Y improved outcomes in dogs with condition X.",
        strata={"species": ["Canine"], "study_design": "Retrospective", "clinical_topic": "Cardiology",
                "journal": "JVIM", "input_source": "processed"},
        requires_human_review=False,
        llm_jury_score=4.2,
        llm_jury_score_weighted=4.3,
        llm_jury_score_unweighted=4.1,
        llm_criteria_scores={"faithfulness": {"score": 4, "reasoning": "ok"}},
    )


def test_packet_markdown_contains_no_model_identifiers() -> None:
    """The rendered reading packet (render_packet_markdown) must never contain
    any of the forbidden AI-identity words (BLIND_FORBIDDEN_TOKENS, defined
    in evaluator.py) — checked by lower-casing the whole rendered text and
    searching for each forbidden word inside it.
    """
    items = [_make_item("item_001"), _make_item("item_002", summarizer="gemini", judge="openai")]
    rendered = render_packet_markdown(items, reviewer_id=1).lower()
    for forbidden in BLIND_FORBIDDEN_TOKENS:
        assert forbidden not in rendered, (
            f"Reviewer packet leaked the token '{forbidden}'. The blind protocol "
            "requires reviewer-facing files to be model-agnostic."
        )


def test_scoresheet_csv_contains_no_model_identifiers() -> None:
    """Same blind check as above, but for the CSV scoresheet instead of the
    reading packet.
    """
    items = [_make_item("item_001"), _make_item("item_002", summarizer="anthropic")]
    rendered = render_scoresheet_csv(items).lower()
    for forbidden in BLIND_FORBIDDEN_TOKENS:
        assert forbidden not in rendered


def test_scoresheet_csv_has_expected_columns_and_blank_scores() -> None:
    """Checks the CSV scoresheet's shape: the right column headers, in the
    right order, one row per item, with the reference columns (item_id,
    article_title) already filled in and every score column left blank for
    the reviewer to type into.
    """
    items = [_make_item("item_001", title="A Retrospective Study of Condition X in Dogs")]
    rendered = render_scoresheet_csv(items)
    reader = csv.DictReader(io.StringIO(rendered))
    assert reader.fieldnames == SCORESHEET_FIELDS
    rows = list(reader)
    assert len(rows) == 1
    # item_id and article_title are pre-filled reference cells; everything the
    # reviewer scores starts blank.
    assert rows[0]["item_id"] == "item_001"
    assert rows[0]["article_title"] == "A Retrospective Study of Condition X in Dogs"
    for field in SCORESHEET_FIELDS:
        if field in ("item_id", "article_title"):
            continue
        assert rows[0][field] == ""


def test_render_reviewer_guide_markdown_returns_the_checked_in_doc() -> None:
    """render_reviewer_guide_markdown() should return the guide document
    exactly as it's saved on disk (docs/booklet/07_human_validation_guide.md),
    unmodified.
    """
    content = render_reviewer_guide_markdown()
    assert "A Guide for Veterinarian Reviewers" in content
    assert content == REVIEWER_GUIDE_PATH.read_text(encoding="utf-8")


def test_render_reviewer_guide_markdown_contains_no_model_identifiers() -> None:
    """The reviewer guide itself must also stay blind — no AI names anywhere
    in it, same check as the packet/scoresheet tests above.
    """
    content = render_reviewer_guide_markdown().lower()
    for forbidden in BLIND_FORBIDDEN_TOKENS:
        assert forbidden not in content, (
            f"Reviewer guide leaked the token '{forbidden}'; the guide must stay "
            "model-agnostic like packet.md and the scoresheet."
        )


def test_render_reviewer_guide_markdown_raises_clearly_when_missing(monkeypatch, tmp_path: Path) -> None:
    """If the guide file is missing from disk, the function should fail with
    a clear, expected error (FileNotFoundError) rather than doing something
    confusing. ``monkeypatch`` is a pytest tool that temporarily swaps out a
    value (here, REVIEWER_GUIDE_PATH) for the duration of one test, then
    automatically restores it afterwards — so this test can point the path
    at a file that doesn't exist without affecting any other test.
    ``tmp_path`` is a pytest-provided empty temporary folder, unique to this
    test run, that gets cleaned up automatically.
    """
    import pytest
    monkeypatch.setattr(human_review, "REVIEWER_GUIDE_PATH", tmp_path / "does_not_exist.md")
    with pytest.raises(FileNotFoundError):
        human_review.render_reviewer_guide_markdown()


def test_unblinding_key_carries_identity_and_llm_scores() -> None:
    """build_unblinding_key() should record, for each item, the real
    identity (doi, summarizer, judge) and the AI jury's own score — this is
    the one place that information is allowed to be written down.
    """
    items = [_make_item("item_001")]
    key = build_unblinding_key(items, seed=42, sample_size=1, reviewer_count=2)
    assert key["items"]["item_001"]["doi"] == "10.1111/test.review"
    assert key["items"]["item_001"]["summarizer"] == "openai"
    assert key["items"]["item_001"]["judge"] == "anthropic"
    assert key["items"]["item_001"]["llm_jury_score"] == 4.2
    assert key["reviewer_count"] == 2
    assert key["seed"] == 42


# ---------------------------------------------------------------------------
# Sampler
# ---------------------------------------------------------------------------
# In plain English: these tests check sample_rows_for_review() — the
# function that decides which already-judged (article, AI summary) rows get
# sent out for human review. They check the three promises that function
# makes: flagged rows come first, the sample spans multiple journals rather
# than clustering in one, and the same seed always reproduces the same pick.

def _row(doi: str, summarizer: str, *, journal: str, requires_review: bool = False,
        timestamp: str = "2026-01-01T00:00:00+00:00", input_source: str = "processed") -> dict:
    """Helper (not a test) that builds one fake evaluations.jsonl row with
    just enough fields for the sampler tests below to exercise
    sample_rows_for_review() without needing a real evaluation run.
    """
    return {
        "doi": doi,
        "summarizer": summarizer,
        "input_source": input_source,
        "requires_human_review": requires_review,
        "timestamp": timestamp,
        "strata": {
            "species": ["Canine"], "study_design": "RCT", "clinical_topic": "Oncology",
            "journal": journal, "input_source": input_source,
        },
    }


def test_sampler_prioritizes_flagged_rows() -> None:
    """3 rows are pre-flagged as requires_human_review, mixed in with 10
    that aren't; asking for a sample of exactly 3 should return only the
    flagged ones.
    """
    rows = [_row(f"10.1/{i}", "openai", journal="JVIM", requires_review=True) for i in range(3)]
    rows += [_row(f"10.1/other{i}", "openai", journal="JVIM") for i in range(10)]
    sampled = sample_rows_for_review(rows, sample_size=3, seed=42)
    assert len(sampled) == 3
    assert all(r["requires_human_review"] for r in sampled)


def test_sampler_fills_remaining_slots_after_flagged() -> None:
    """With only 1 flagged row but a sample size of 4, the flagged row is
    included plus 3 more from the unflagged remainder.
    """
    flagged = [_row("10.1/flag0", "openai", journal="JVIM", requires_review=True)]
    remainder = [_row(f"10.1/r{i}", "openai", journal="JVIM") for i in range(10)]
    sampled = sample_rows_for_review(flagged + remainder, sample_size=4, seed=42)
    assert len(sampled) == 4
    assert sum(1 for r in sampled if r["requires_human_review"]) == 1


def test_sampler_stratifies_across_journals() -> None:
    """With 3 journals contributing 10 rows each, a sample of 6 should touch
    all 3 journals rather than, by bad luck, landing entirely within one.
    """
    rows = []
    for journal in ("JVIM", "AJVR", "VRU"):
        rows += [_row(f"10.1/{journal}_{i}", "openai", journal=journal) for i in range(10)]
    sampled = sample_rows_for_review(rows, sample_size=6, seed=42)
    journals_seen = {r["strata"]["journal"] for r in sampled}
    assert len(journals_seen) == 3, (
        "A stratified sample of 6 across 3 evenly-sized journal groups should "
        f"span all three journals, saw only {journals_seen}"
    )


def test_sampler_is_deterministic_for_a_fixed_seed() -> None:
    """Running the sampler twice with the exact same seed must produce the
    exact same sample, in the exact same order — this is what makes a lost
    reviewer sheet regenerable.
    """
    rows = []
    for journal in ("JVIM", "AJVR"):
        rows += [_row(f"10.1/{journal}_{i}", "openai", journal=journal) for i in range(8)]
    first = sample_rows_for_review(rows, sample_size=5, seed=7)
    second = sample_rows_for_review(rows, sample_size=5, seed=7)
    assert [r["doi"] for r in first] == [r["doi"] for r in second]


def test_sampler_dedupes_to_latest_row_per_item() -> None:
    """Two rows share the same (doi, summarizer) but different timestamps;
    the sampler should keep only the newer one (see human_review._dedupe_rows).
    """
    older = _row("10.1/dup", "openai", journal="JVIM", timestamp="2026-01-01T00:00:00+00:00")
    older["requires_human_review"] = False
    newer = _row("10.1/dup", "openai", journal="JVIM", timestamp="2026-02-01T00:00:00+00:00")
    newer["requires_human_review"] = True
    sampled = sample_rows_for_review([older, newer], sample_size=5, seed=1)
    assert len(sampled) == 1
    assert sampled[0]["requires_human_review"] is True


def test_sampler_returns_empty_for_zero_sample_size() -> None:
    """Asking for zero items should return an empty list, not an error."""
    rows = [_row("10.1/a", "openai", journal="JVIM")]
    assert sample_rows_for_review(rows, sample_size=0, seed=1) == []


def test_sampler_default_sample_unit_matches_explicit_items() -> None:
    """Not passing sample_unit at all should behave identically to passing
    sample_unit="items" explicitly (the low-level sampler's own default).
    """
    rows = [_row(f"10.1/{i}", "openai", journal="JVIM") for i in range(5)]
    default = sample_rows_for_review(rows, sample_size=3, seed=42)
    explicit = sample_rows_for_review(rows, sample_size=3, seed=42, sample_unit="items")
    assert default == explicit


def test_sample_rows_for_review_rejects_unknown_sample_unit() -> None:
    """Passing a made-up sample_unit value should fail loudly with a
    ValueError instead of doing something unexpected.
    """
    import pytest
    rows = [_row("10.1/a", "openai", journal="JVIM")]
    with pytest.raises(ValueError):
        sample_rows_for_review(rows, sample_size=1, seed=1, sample_unit="bogus")


# ---------------------------------------------------------------------------
# Sampler — sample_unit="articles" (read once, score every provider)
# ---------------------------------------------------------------------------
# In plain English: these tests cover the OTHER sampling mode, where
# sample_size counts articles rather than individual rows, and every
# provider's summary of a chosen article rides along with it — the shape
# the real export actually uses by default.

def _multi_provider_rows(num_articles: int, *, journal: str = "JVIM",
                          providers: tuple[str, ...] = ("openai", "anthropic", "gemini")) -> list[dict]:
    """Helper (not a test): builds fake rows for N articles, each with all 3
    providers' summaries present — the shape "articles" mode is designed
    for.
    """
    rows = []
    for i in range(num_articles):
        doi = f"10.1/article{i}"
        rows += [_row(doi, provider, journal=journal) for provider in providers]
    return rows


def test_sampler_articles_mode_includes_every_provider_of_each_selected_article() -> None:
    """Sampling 2 articles in "articles" mode should return 6 rows total (2
    articles x 3 providers each) — every provider's summary of each chosen
    article, not just one.
    """
    rows = _multi_provider_rows(5)
    sampled = sample_rows_for_review(rows, sample_size=2, seed=42, sample_unit="articles")
    dois_seen = {r["doi"] for r in sampled}
    assert len(dois_seen) == 2
    assert len(sampled) == 6  # 2 articles x 3 providers
    for doi in dois_seen:
        summarizers = {r["summarizer"] for r in sampled if r["doi"] == doi}
        assert summarizers == {"openai", "anthropic", "gemini"}


def test_sampler_articles_mode_interleaves_sibling_rows_apart() -> None:
    """Two rows for the same article (different providers' summaries) must
    never end up next to each other in the sampled order — see
    human_review._interleave_article_groups.
    """
    rows = _multi_provider_rows(5)
    sampled = sample_rows_for_review(rows, sample_size=5, seed=42, sample_unit="articles")
    assert len(sampled) == 15
    for idx in range(len(sampled) - 1):
        assert sampled[idx]["doi"] != sampled[idx + 1]["doi"], (
            f"Sibling rows for {sampled[idx]['doi']} landed adjacent at index {idx}; "
            "articles mode must interleave same-article rows apart so a reviewer "
            "never scores two summaries of the same article back to back."
        )


def test_sampler_articles_mode_prioritizes_flagged_articles() -> None:
    """An article counts as "flagged" if ANY of its provider rows is flagged
    — here only the openai row is flagged, but the whole article (all 3
    provider rows) should still be picked first.
    """
    flagged_doi = "10.1/flagged"
    rows = [_row(flagged_doi, p, journal="JVIM", requires_review=(p == "openai"))
            for p in ("openai", "anthropic", "gemini")]
    rows += _multi_provider_rows(5, journal="JVIM")
    sampled = sample_rows_for_review(rows, sample_size=1, seed=1, sample_unit="articles")
    assert {r["doi"] for r in sampled} == {flagged_doi}
    assert len(sampled) == 3


def test_sampler_articles_mode_stratifies_across_journals() -> None:
    """Same stratification guarantee as the "items" mode test above, checked
    here at article granularity instead of row granularity.
    """
    rows = []
    for journal in ("JVIM", "AJVR", "VRU"):
        for i in range(10):
            doi = f"10.1/{journal}_article{i}"
            rows += [_row(doi, provider, journal=journal)
                     for provider in ("openai", "anthropic", "gemini")]
    sampled = sample_rows_for_review(rows, sample_size=6, seed=42, sample_unit="articles")
    journals_seen = {r["strata"]["journal"] for r in sampled}
    assert len(journals_seen) == 3, (
        "A stratified sample of 6 articles across 3 evenly-sized journal groups "
        f"should span all three journals, saw only {journals_seen}"
    )


def test_sampler_articles_mode_is_deterministic_for_a_fixed_seed() -> None:
    """Same reproducibility guarantee as "items" mode, checked here for
    "articles" mode.
    """
    rows = _multi_provider_rows(8)
    first = sample_rows_for_review(rows, sample_size=4, seed=7, sample_unit="articles")
    second = sample_rows_for_review(rows, sample_size=4, seed=7, sample_unit="articles")
    assert [(r["doi"], r["summarizer"]) for r in first] == [(r["doi"], r["summarizer"]) for r in second]


# ---------------------------------------------------------------------------
# Full export (offline, mock corpus)
# ---------------------------------------------------------------------------
# In plain English: these tests exercise the WHOLE export pipeline end to
# end (export_human_review), not just one small piece of it — building a
# tiny fake "corpus" (a fake evaluations.jsonl + fake article/summary text)
# on disk in a temporary folder, running the real export function against
# it, and checking the folders/files it produces. "offline, mock corpus"
# means no real data or network calls are involved, just made-up test data.

# The real export always draws EXACTLY one article per one of the study's 5
# journals per 5-article quota, so a fixture corpus for export_human_review
# must cover all 5 -- this is that fixed DOI-per-journal mapping, reused by
# every test below that needs a full 5-article (sample_size=5) draw.
_FIXTURE_DOIS = [f"10.1111/test.{i}" for i in range(len(STUDY_JOURNALS))]


def _write_fixture_corpus(tmp_path: Path, monkeypatch) -> tuple[Path, Path, Path]:
    """Build a minimal evaluations.jsonl + summaries.jsonl/processed cache pair.

    One DOI per study journal (_FIXTURE_DOIS / STUDY_JOURNALS), one
    summarizer slot each, all successfully judged and joinable back to real
    reference/candidate text via prepare_texts.PROCESSED_DIR (patched here
    exactly as tests/test_eval_instances.py does). Also writes a matching
    dummy PDF per article under a returned ``raw_dir``, so export's default
    path (copy the real source PDF into ``article.pdf``) is exercised rather
    than only the ``article.md`` fallback.

    In plain English: this helper fakes an entire tiny "study" — one article
    per journal, each already summarized (by "openai") and already judged,
    with its "original PDF" sitting on disk too — and writes it to temporary
    files, so the tests below can run the real export function (which always
    samples evenly across all 5 real study journals) against realistic-
    looking input without needing the actual multi-gigabyte project data.
    ``tmp_path`` is pytest's built-in "give me an empty scratch folder for
    this test" fixture.
    """
    processed_dir = tmp_path / "processed"
    processed_dir.mkdir()
    monkeypatch.setattr(prepare_texts, "PROCESSED_DIR", processed_dir)

    dois = _FIXTURE_DOIS
    summaries_path = tmp_path / "summaries.jsonl"
    evaluations_path = tmp_path / "evaluations.jsonl"
    raw_dir = tmp_path / "raw"
    raw_dir.mkdir()

    with open(summaries_path, "w", encoding="utf-8") as f:
        for doi, journal in zip(dois, STUDY_JOURNALS):
            f.write(json.dumps({
                "doi": doi,
                "journal": journal,
                "models": {"openai": {"status": "success", "summary": f"Summary of {doi}."}},
            }) + "\n")

    for doi in dois:
        slug = doi_to_slug(doi)
        (processed_dir / f"{slug}.jsonl").write_text(
            json.dumps({"doi": doi, "slug": slug, "text": f"Full cleaned article text for {doi}."}) + "\n",
            encoding="utf-8",
        )

    for doi, journal in zip(dois, STUDY_JOURNALS):
        record = {"doi": doi, "journal": journal}
        (raw_dir / descriptive_pdf_filename(record)).write_bytes(b"%PDF-1.4 dummy\n")

    with open(evaluations_path, "w", encoding="utf-8") as f:
        for i, (doi, journal) in enumerate(zip(dois, STUDY_JOURNALS)):
            f.write(json.dumps({
                "doi": doi,
                "summarizer": "openai",
                "judge": "anthropic",
                "input_source": "processed",
                "rubric_version": "vet_medhelm_score_v1.0",
                "jury_score": 4.0,
                "jury_score_weighted": 4.0,
                "jury_score_unweighted": 4.0,
                "criteria_scores": {"faithfulness": {"score": 4, "reasoning": "ok"}},
                "requires_human_review": i == 0,
                "strata": {"species": ["Canine"], "study_design": "RCT", "clinical_topic": "Cardiology",
                          "journal": journal, "input_source": "processed"},
                "timestamp": "2026-01-01T00:00:00+00:00",
            }) + "\n")

    return summaries_path, evaluations_path, raw_dir


def test_export_human_review_creates_reviewer_folders(tmp_path: Path, monkeypatch) -> None:
    """The big end-to-end check: two consecutive export_human_review() calls
    (overlap_ratio=1.0, so human2 repeats human1's items) over the 5-journal
    fixture corpus should each produce a complete, correctly structured
    humanN/ folder (guide + packet + xlsx scoresheet + per-item
    article.pdf/summary files), a matching sibling unblinding key, and —
    critically — none of it anywhere containing a forbidden AI-identity word.
    """
    summaries_path, evaluations_path, raw_dir = _write_fixture_corpus(tmp_path, monkeypatch)
    output_dir = tmp_path / "human_review"

    kwargs = dict(
        sample_size=len(STUDY_JOURNALS), overlap_ratio=1.0, seed=42,
        evaluations_path=evaluations_path, output_dir=output_dir,
        summaries_path=summaries_path,
        summaries_txt_dir=tmp_path / "no_summaries_txt",
        dev_tests_summaries_txt_dir=tmp_path / "no_dev_summaries_txt",
        raw_dir=raw_dir,
    )
    results = [export_human_review(**kwargs) for _ in range(2)]

    assert [r.human_number for r in results] == [1, 2]
    for result in results:
        assert result.items_exported == len(STUDY_JOURNALS)
        assert result.skipped_rows == 0

    expected_guide = REVIEWER_GUIDE_PATH.read_text(encoding="utf-8")
    last_item_id = f"item_{len(STUDY_JOURNALS):03d}"
    for human_number in (1, 2):
        human_dir = output_dir / f"human{human_number}"
        guide = (human_dir / "REVIEWER_GUIDE.md").read_text(encoding="utf-8")
        packet = (human_dir / "packet.md").read_text(encoding="utf-8")
        # Scoresheet is now .xlsx (not .csv); the packet is a navigation index.
        assert (human_dir / f"scoresheet_human{human_number}.xlsx").exists()
        assert guide == expected_guide
        assert "item_001" in packet and last_item_id in packet
        assert f"scoresheet_human{human_number}.xlsx" in packet
        assert "article.pdf" in packet
        # The real source PDF + candidate summary live in per-item folders now
        # (article.pdf is a byte-for-byte copy — a real PDF is found for every
        # fixture DOI, so no item falls back to the article.md text dump).
        item_dir = human_dir / "item_001"
        article_pdf = item_dir / "article.pdf"
        summary = (item_dir / "summary.md").read_text(encoding="utf-8")
        assert article_pdf.exists()
        assert not (item_dir / "article.md").exists()
        assert article_pdf.read_bytes() == b"%PDF-1.4 dummy\n"
        assert not (human_dir / "original_articles").exists()
        blind_text = "\n".join([
            guide, packet, article_pdf.read_bytes().decode("ascii", errors="ignore"), summary,
        ]).lower()
        for forbidden in BLIND_FORBIDDEN_TOKENS:
            assert forbidden not in blind_text

        key = json.loads(
            (output_dir / f"unblinding_key_human{human_number}.json").read_text(encoding="utf-8")
        )
        assert set(key["items"]) == {f"item_{i:03d}" for i in range(1, len(STUDY_JOURNALS) + 1)}
        assert {item["doi"] for item in key["items"].values()} == set(_FIXTURE_DOIS)
        assert all(item["summarizer"] == "openai" for item in key["items"].values())

    # overlap_ratio=1.0 -> human2 should carry every article human1 saw.
    assert results[1].overlap_units == len(STUDY_JOURNALS)


def test_export_human_review_falls_back_to_article_md_without_source_pdf(
    tmp_path: Path, monkeypatch,
) -> None:
    """When no source PDF can be resolved for a DOI, the item folder falls
    back to article.md (the cached text the AI summarizer actually read)
    instead of being left without an article at all — and the export warns
    about every DOI that fell back.
    """
    summaries_path, evaluations_path, _real_raw_dir = _write_fixture_corpus(tmp_path, monkeypatch)
    empty_raw_dir = tmp_path / "empty_raw"  # deliberately has no matching PDFs
    empty_raw_dir.mkdir()
    output_dir = tmp_path / "human_review"

    result = export_human_review(
        sample_size=len(STUDY_JOURNALS),
        evaluations_path=evaluations_path,
        output_dir=output_dir,
        summaries_path=summaries_path,
        summaries_txt_dir=tmp_path / "no_summaries_txt",
        dev_tests_summaries_txt_dir=tmp_path / "no_dev_summaries_txt",
        raw_dir=empty_raw_dir,
    )

    human_dir = output_dir / f"human{result.human_number}"
    for item_dir in sorted(p for p in human_dir.glob("item_*") if p.is_dir()):
        assert not (item_dir / "article.pdf").exists()
        article = (item_dir / "article.md").read_text(encoding="utf-8")
        assert "Full cleaned article text for" in article


def _write_fixture_corpus_multi_provider(
    tmp_path: Path, monkeypatch, *, num_articles: int = len(STUDY_JOURNALS),
) -> tuple[Path, Path]:
    """Like _write_fixture_corpus, but every article has all 3 providers
    successfully summarized and evaluated -- used to test the exact
    per-journal quota sampler with multiple providers per article.

    One article per study journal (cycling through STUDY_JOURNALS if
    ``num_articles`` exceeds it), each with THREE AI summaries (openai,
    anthropic, gemini) -- what's needed to meaningfully test "every
    provider's summary of a sampled article is included".
    """
    processed_dir = tmp_path / "processed_multi"
    processed_dir.mkdir()
    monkeypatch.setattr(prepare_texts, "PROCESSED_DIR", processed_dir)

    providers = ("openai", "anthropic", "gemini")
    dois = [f"10.1111/test.multi{i}" for i in range(num_articles)]
    journals = [STUDY_JOURNALS[i % len(STUDY_JOURNALS)] for i in range(num_articles)]
    summaries_path = tmp_path / "summaries_multi.jsonl"
    evaluations_path = tmp_path / "evaluations_multi.jsonl"

    with open(summaries_path, "w", encoding="utf-8") as f:
        for doi, journal in zip(dois, journals):
            f.write(json.dumps({
                "doi": doi,
                "journal": journal,
                "models": {p: {"status": "success", "summary": f"Summary variant {i} of {doi}."}
                           for i, p in enumerate(providers)},
            }) + "\n")

    for doi in dois:
        slug = doi_to_slug(doi)
        (processed_dir / f"{slug}.jsonl").write_text(
            json.dumps({"doi": doi, "slug": slug, "text": f"Full cleaned article text for {doi}."}) + "\n",
            encoding="utf-8",
        )

    with open(evaluations_path, "w", encoding="utf-8") as f:
        for doi, journal in zip(dois, journals):
            for provider in providers:
                f.write(json.dumps({
                    "doi": doi,
                    "summarizer": provider,
                    "judge": "anthropic",
                    "input_source": "processed",
                    "rubric_version": "vet_medhelm_score_v1.0",
                    "jury_score": 4.0,
                    "jury_score_weighted": 4.0,
                    "jury_score_unweighted": 4.0,
                    "criteria_scores": {"faithfulness": {"score": 4, "reasoning": "ok"}},
                    "requires_human_review": False,
                    "strata": {"species": ["Canine"], "study_design": "RCT", "clinical_topic": "Cardiology",
                              "journal": journal, "input_source": "processed"},
                    "timestamp": "2026-01-01T00:00:00+00:00",
                }) + "\n")

    return summaries_path, evaluations_path


def test_export_human_review_articles_mode_reads_each_article_once(tmp_path: Path, monkeypatch) -> None:
    """sample_size=5 (1 article/journal) should read exactly 5 articles but
    export all 3 providers' summaries of each (15 items total).

    In plain English: this is the full export pipeline test for the exact
    per-journal quota, checking that the resulting unblinding key really
    does show 5 distinct articles (one per journal), each paired with all 3
    providers, and that the written packet still passes the blind check.
    """
    summaries_path, evaluations_path = _write_fixture_corpus_multi_provider(tmp_path, monkeypatch)
    output_dir = tmp_path / "human_review_articles"

    result = export_human_review(
        sample_size=len(STUDY_JOURNALS),
        seed=42,
        evaluations_path=evaluations_path,
        output_dir=output_dir,
        summaries_path=summaries_path,
        summaries_txt_dir=tmp_path / "no_summaries_txt",
        dev_tests_summaries_txt_dir=tmp_path / "no_dev_summaries_txt",
    )

    assert result.items_exported == len(STUDY_JOURNALS) * 3  # N articles x 3 providers

    key = json.loads(
        (output_dir / f"unblinding_key_human{result.human_number}.json").read_text(encoding="utf-8")
    )
    dois = {item["doi"] for item in key["items"].values()}
    assert len(dois) == len(STUDY_JOURNALS)
    for doi in dois:
        summarizers = {item["summarizer"] for item in key["items"].values() if item["doi"] == doi}
        assert summarizers == {"openai", "anthropic", "gemini"}

    human_dir = output_dir / f"human{result.human_number}"
    assert (human_dir / "REVIEWER_GUIDE.md").exists()
    packet = (human_dir / "packet.md").read_text(encoding="utf-8")
    for forbidden in BLIND_FORBIDDEN_TOKENS:
        assert forbidden not in packet.lower()


def test_export_human_review_skips_rows_missing_source_text(tmp_path: Path, monkeypatch) -> None:
    """A row whose (doi, summarizer, input_source) no longer resolves to a
    live instance (source text moved/regenerated) is skipped, not fatal.

    In plain English: simulates the case where evaluations.jsonl mentions an
    article whose text can no longer be found (e.g. the underlying data
    changed since the article was judged) — export should quietly skip that
    one item and continue, rather than crashing the whole export.
    """
    summaries_path, evaluations_path, _raw_dir = _write_fixture_corpus(tmp_path, monkeypatch)

    # Add an extra evaluation row for a DOI with no matching summaries.jsonl
    # entry, in the SAME journal as one of the 5 real fixture articles
    # (STUDY_JOURNALS[0]) and flagged requires_human_review=True so the
    # flagged-first selection deterministically picks it over that journal's
    # real (resolvable) article for the 1/journal quota.
    with open(evaluations_path, "a", encoding="utf-8") as f:
        f.write(json.dumps({
            "doi": "10.1111/test.orphan",
            "summarizer": "openai",
            "judge": "anthropic",
            "input_source": "processed",
            "rubric_version": "vet_medhelm_score_v1.0",
            "jury_score": 3.0,
            "requires_human_review": True,
            "strata": {"journal": STUDY_JOURNALS[0], "input_source": "processed"},
            "timestamp": "2026-01-01T00:00:00+00:00",
        }) + "\n")

    result = export_human_review(
        sample_size=len(STUDY_JOURNALS),
        seed=42,
        evaluations_path=evaluations_path,
        output_dir=tmp_path / "human_review",
        summaries_path=summaries_path,
        summaries_txt_dir=tmp_path / "no_summaries_txt",
        dev_tests_summaries_txt_dir=tmp_path / "no_dev_summaries_txt",
    )

    # The orphan's journal now has 2 candidates (real + orphan); the other 4
    # journals have exactly 1 each -> 5 articles selected total, but only 4
    # resolve to real source text (the orphan is skipped, not fatal).
    assert result.items_exported == len(STUDY_JOURNALS) - 1
    assert result.skipped_rows == 1


def test_export_human_review_rejects_nonpositive_or_uneven_sample_size(tmp_path: Path) -> None:
    """A zero/negative sample size, or one that doesn't divide evenly across
    the 5 study journals, should be rejected immediately with a ValueError
    rather than silently producing an empty or unevenly-sampled export.
    """
    import pytest
    with pytest.raises(ValueError):
        export_human_review(sample_size=0, evaluations_path=tmp_path / "evaluations.jsonl")
    with pytest.raises(ValueError):
        export_human_review(sample_size=-1, evaluations_path=tmp_path / "evaluations.jsonl")
    with pytest.raises(ValueError):
        # 7 is not a multiple of len(STUDY_JOURNALS) == 5.
        export_human_review(sample_size=7, evaluations_path=tmp_path / "evaluations.jsonl")


def test_cli_main_reports_failure_when_nothing_to_export(tmp_path: Path) -> None:
    """Running the command-line export against an empty (nonexistent)
    evaluations.jsonl should exit with code 1 (failure) rather than
    pretending everything worked.
    """
    empty_evaluations = tmp_path / "evaluations.jsonl"
    exit_code = human_review.main([
        "--evaluations", str(empty_evaluations),
        "--output-dir", str(tmp_path / "human_review"),
    ])
    assert exit_code == 1


# ---------------------------------------------------------------------------
# Interactive article-count prompt
# ---------------------------------------------------------------------------
# In plain English: these tests check prompt_article_count() — the function
# that, when run interactively, asks "how many articles will you evaluate?"
# and keeps re-asking until it gets a sensible answer.

class _FakeTTY(io.StringIO):
    """A stand-in for a real interactive terminal. io.StringIO is normally
    an in-memory, non-interactive text stream (its isatty() reports False);
    this small subclass overrides isatty() to always report True, so
    prompt_article_count() believes it's talking to a real person typing
    answers and exercises its "ask, validate, re-ask" loop.
    """
    def isatty(self) -> bool:  # noqa: D401 - a fake interactive stream
        return True


def test_prompt_article_count_reasks_below_minimum_and_on_non_numeric() -> None:
    """Feeding the prompt a too-low answer ("2"), then a non-numeric one
    ("not-a-number"), then a valid one ("7") should make it re-ask twice and
    finally return 7 — i.e. it never accepts a bad answer.
    """
    answers = iter(["2", "not-a-number", "7"])
    got = human_review.prompt_article_count(
        human_review.MIN_ARTICLES,
        input_fn=lambda _prompt="": next(answers),
        stream=_FakeTTY(),
    )
    assert got == 7


def test_prompt_article_count_falls_back_when_not_a_tty() -> None:
    # A non-interactive stream (io.StringIO.isatty() is False) must not block;
    # it falls back to max(default, MIN_ARTICLES).
    assert human_review.prompt_article_count(3, stream=io.StringIO()) == human_review.MIN_ARTICLES
    assert human_review.prompt_article_count(8, stream=io.StringIO()) == 8


# ---------------------------------------------------------------------------
# Ingest reads the .xlsx scoresheet the export now writes
# ---------------------------------------------------------------------------

def test_ingest_reads_filled_xlsx_scoresheet(tmp_path: Path, monkeypatch) -> None:
    """A small end-to-end round trip: export a real .xlsx scoresheet, fill
    it in programmatically (simulating a reviewer typing scores into Excel),
    then ingest it and check the resulting human_reviews.jsonl row has the
    scores and comment that were "typed in".
    """
    from openpyxl import load_workbook

    summaries_path, evaluations_path, _raw_dir = _write_fixture_corpus(tmp_path, monkeypatch)
    output_dir = tmp_path / "human_review"
    export_human_review(
        sample_size=len(STUDY_JOURNALS),
        seed=42,
        evaluations_path=evaluations_path,
        output_dir=output_dir,
        summaries_path=summaries_path,
        summaries_txt_dir=tmp_path / "no_summaries_txt",
        dev_tests_summaries_txt_dir=tmp_path / "no_dev_summaries_txt",
    )

    sheet = output_dir / "human1" / "scoresheet_human1.xlsx"
    assert sheet.exists()

    col = {f: SCORESHEET_FIELDS.index(f) + 1 for f in SCORESHEET_FIELDS}
    wb = load_workbook(sheet)
    ws = wb.active
    # Fill item_001 (row 2) with a full set of scores + a hallucination flag.
    for criterion in human_review.CRITERIA:
        ws.cell(row=2, column=col[criterion], value=4)
    ws.cell(row=2, column=col["hallucination_present"], value="no")
    ws.cell(row=2, column=col["comment"], value="reads well")
    wb.save(sheet)

    output_path = tmp_path / "human_reviews.jsonl"
    result = human_review.ingest_human_reviews(review_dir=output_dir, output_path=output_path)
    assert result.rows_written == 1

    rows = [json.loads(line) for line in output_path.read_text(encoding="utf-8").splitlines() if line]
    assert len(rows) == 1
    assert rows[0]["item_id"] == "item_001"
    assert rows[0]["reviewer_id"] == "human1"
    assert rows[0]["criteria_scores"]["faithfulness"] == 4.0
    assert rows[0]["hallucination_present"] is False
    assert rows[0]["comment"] == "reads well"
