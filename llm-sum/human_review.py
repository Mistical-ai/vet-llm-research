"""
llm-sum/human_review.py — Human validation: stratified sampling + blind export
================================================================================

IN PLAIN ENGLISH
-----------------
This file has two jobs, done in order:

1. EXPORT — pick a fair sample of articles, and for each one, package up the
   original article text plus one AI-written summary into a folder a real
   veterinarian can read and score by hand — WITHOUT telling them which AI
   (OpenAI, Anthropic, Gemini, ...) wrote that summary. This "don't reveal
   who wrote it" rule is called the blind protocol, and it's the same rule
   the AI judges themselves follow in ``evaluator.py``. The only place the
   AI's identity is written down is a private per-reviewer key file
   (``unblinding_key_human{N}.json``, one per ``humanN/`` folder), which
   reviewers never see.
2. INGEST — once a reviewer has filled in their scoring spreadsheet, read it
   back in, look up the private key to find out which AI actually wrote each
   summary, and save everything in one standard file
   (``data/human_reviews.jsonl``) so later code can compare "what the human
   said" against "what the AI judges said" and see whether they agree.

Think of it like a blind taste test: tasters (reviewers) are given unlabeled
samples (article + summary) and asked to rate them, and only afterwards does
someone check the answer key to see which brand (AI) each sample was.

WHY THIS MODULE EXISTS
-----------------------
The blind LLM jury in ``evaluator.py`` is the study's authoritative score, but
a jury score with no check against expert judgment is hard to defend in a
methods section. This module samples already-judged (paper, summariser) pairs
from ``data/evaluations.jsonl`` and exports them as blind review packets a
veterinarian can score by hand — one or several independent reviewers, no
code required on their end.

THE BLIND PROTOCOL APPLIES HERE TOO
------------------------------------
Exactly like ``evaluator.build_judge_prompt()``, the reviewer-facing files
this module writes carry only ``item_id`` + the original article text + the
candidate summary — never the summariser name, the judge's score, or any
other clue to identity. The mapping from ``item_id`` back to
(doi, summariser, judge, LLM scores) lives in a single un-blinding key file
that is never handed to a reviewer (see ``build_unblinding_key``).

INGEST + ANALYSIS (Phase 5, Chunk 6)
------------------------------------
Once reviewers return their filled scoresheets, ``ingest_human_reviews()``
re-joins each ``item_id`` to the private un-blinding key, normalizes the rows
into ``data/human_reviews.jsonl``, and ``analyze_human_reviews()`` answers the
two questions that make the LLM jury defensible:

- *Do the human reviewers agree with each other?* (inter-reviewer
  Krippendorff's alpha — reuses ``reliability.compute_reliability``, treating
  each reviewer as a "rater" exactly as it treats each judge; skipped for a
  lone reviewer, who has no one to agree with.)
- *Does the LLM jury track expert judgment?* (human-vs-jury Pearson +
  Spearman correlation and a Bland-Altman bias, overall and per-criterion —
  reported even for the 1-reviewer case.)

DESIGN CONSTRAINTS (from CLAUDE.md)
-------------------------------------
- Offline only: reads ``data/evaluations.jsonl`` + already-cached summary
  text; never calls a paid API.
- Dependency-light: only the standard library directly (csv, json, random);
  correlation statistics come via reliability.py (which uses scipy.stats).
- Reuses existing joins (``eval_instances.iter_evaluation_instances``,
  ``summarize_all_ingest.iter_summarize_all_instances``,
  ``eval_report.iter_evaluation_rows``) instead of re-deriving them.
"""

# This line just tells Python "treat type hints loosely" (a housekeeping
# detail for newer type-hint syntax like ``str | None``); it doesn't change
# what the program does.
from __future__ import annotations

from _bootstrap import *  # noqa: F401,F403

# Everything below is a "standard library" import — built-in Python tools
# that ship with Python itself, no separate install needed:
#   argparse  — reads command-line options like --overlap-ratio (see main()).
#   csv       — reads/writes spreadsheet-style comma-separated-value files.
#   io        — lets code build a file's contents in memory before saving it.
#   json      — reads/writes JSON, a simple text format for structured data.
#   os        — reads environment variables / talks to the operating system.
#   random    — controls "randomness" (e.g. which articles get sampled).
#   re        — pattern-matching on text (used for "human1", "human2", ...).
#   shutil    — file utilities such as copying files (used for PDFs).
#   sys       — access to command-line arguments and program exit codes.
import argparse
import csv
import io
import json
import os
import random
import re
import shutil
import sys
# defaultdict: a dictionary that auto-creates a default value (e.g. an empty
# list) the first time a new key is used, so you don't have to check "does
# this key exist yet?" before adding to it.
from collections import defaultdict
# dataclass: a decorator that turns a plain class into a labeled form with
# fixed fields — write the field names once and Python automatically builds
# the boilerplate (constructor, printing, equality) for you. Used below for
# ReviewItem, ExportResult, and IngestResult.
from dataclasses import dataclass
from datetime import datetime, timezone
# Path: an object representing a file or folder location that works the same
# way on Windows and Mac/Linux, instead of hand-building path strings.
from pathlib import Path
# Any: "this could be any type, we're not restricting it." Iterable: "this
# can be looped over with a for-loop" (a list, a generator, etc.). Both are
# type hints — notes for humans and tools about what kind of value a
# function expects or returns; Python does not enforce them at runtime.
from typing import Any, Iterable

# The rest of these imports pull in other files from this same project
# (llm-sum/), reusing code that already exists rather than duplicating it.
from evaluator import (  # noqa: E402  (single source of truth for blind tokens + weights)
    BLIND_FORBIDDEN_TOKENS,
    MEDHELM_CRITERION_WEIGHTS,
    calculate_jury_score,
)
from eval_instances import EvaluationInstance, STRATIFICATION_FIELDS, iter_evaluation_instances  # noqa: E402
from summarize_all_ingest import iter_summarize_all_instances  # noqa: E402
from eval_report import iter_evaluation_rows  # noqa: E402
from file_paths import resolve_existing_pdf_path  # noqa: E402  (DOI/record -> data/raw PDF)
from collect import JOURNAL_TARGETS  # noqa: E402  (single source of truth for the 5 study journals)
import reliability  # noqa: E402  (agreement + correlation stats, reused for humans)
from stats_engine import categorical_agreement  # noqa: E402  (Cohen's Kappa + percent agreement)

# File/folder locations this module reads from or writes to. Each is a Path
# built by joining a name onto DATA_DIR (the project's data/ folder, defined
# in _bootstrap). Keeping them as named constants here means every function
# below can refer to e.g. HUMAN_REVIEW_DIR instead of retyping the path.
EVALUATIONS_PATH = DATA_DIR / "evaluations.jsonl"
SUMMARIES_PATH = DATA_DIR / "summaries.jsonl"
SUMMARIES_TXT_DIR = DATA_DIR / "summaries_txt"
DEV_TESTS_SUMMARIES_TXT_DIR = DATA_DIR / "dev_tests" / "summaries_txt"
RAW_DIR = DATA_DIR / "raw"  # source PDFs copied into each reviewer folder's original_articles/
HUMAN_REVIEW_DIR = DATA_DIR / "human_review"

# In plain English: MIN_ARTICLES is the smallest number of articles allowed
# in a review batch, one per journal (JVIM, JAVMA, Surgery, VRU, JFMS). The
# interactive prompt (``prompt_article_count``) never accepts fewer; each
# article is paired with every provider's summary, so
# 5 articles x 3 providers = 15 scored items (the default export shape).
MIN_ARTICLES = 5

# The 5 study journals, in the same order/spelling as the corpus was
# collected (single source of truth: ``collect.JOURNAL_TARGETS``, the dict
# ``collect.py`` uses to query CrossRef). Reusing it here means this module
# can never drift out of sync with what "journal" values actually appear in
# ``strata.journal`` on an evaluation row.
STUDY_JOURNALS: tuple[str, ...] = tuple(JOURNAL_TARGETS.keys())

# The only three article counts the interactive prompt
# (``prompt_article_count_choice``) offers, each an exact multiple of
# ``len(STUDY_JOURNALS)`` so it always divides evenly into a whole number of
# articles per journal (5 -> 1/journal, 10 -> 2/journal, 25 -> 5/journal),
# each article expanded to ~3 providers' summaries (15 / 30 / 75 to read).
ARTICLE_COUNT_CHOICES: tuple[int, ...] = (5, 10, 25)


class JournalQuotaError(RuntimeError):
    """Raised when the eligible pool can't fill an exact per-journal quota.

    In plain English: this is the error the export raises instead of
    silently handing out an unequal sample when, say, a 10-article request
    (2 articles/journal) is asked of a pool that only has 1 evaluated VRU
    article. ``shortfall`` maps each short journal to
    ``(available, needed)`` so the message can name every short journal at
    once, not just the first one found.
    """

    def __init__(self, shortfall: dict[str, tuple[int, int]]):
        self.shortfall = dict(shortfall)
        details = ", ".join(
            f"{journal} has {available} evaluated article(s) (need {needed})"
            for journal, (available, needed) in sorted(self.shortfall.items())
        )
        super().__init__(
            f"Cannot fill the requested per-journal quota: {details}. Run more "
            "'evaluate' first, or choose a smaller article count "
            f"({', '.join(str(c) for c in ARTICLE_COUNT_CHOICES)})."
        )


# The zero-jargon, standalone reviewer guide (docs/booklet chapter 7). This is
# the single source of truth for reviewer-facing instructions; export copies
# it verbatim into every reviewer folder (see render_reviewer_guide_markdown)
# so a reviewer handed only their humanN/ folder still gets the full guide.
REVIEWER_GUIDE_PATH = REPO_ROOT / "docs" / "booklet" / "07_human_validation_guide.md"
# The normalized ingest output. A derived snapshot (rewritten idempotently on
# every ingest), NOT append-only like evaluations.jsonl — re-ingesting the same
# filled sheets must reproduce the same file, never grow it.
HUMAN_REVIEWS_PATH = DATA_DIR / "human_reviews.jsonl"

# Mirrors evaluator.MEDHELM_CRITERION_WEIGHTS's keys — the same five criteria a
# human reviewer scores, kept as a local tuple so this module has no import-time
# dependency on the evaluator's runtime weight configuration (same rationale as
# reliability.CRITERIA).
CRITERIA = ("faithfulness", "completeness", "clinical_usefulness", "clarity", "safety")
# ``item_id`` + ``article_title`` are pre-filled reference columns (the reviewer
# reads them to find/identify the row, never edits them); everything after is
# what the reviewer fills in. ``article_title`` is ignored by ingest
# (_normalize_scoresheet_row only reads item_id + the scored columns), so adding
# it is purely a reviewer-usability aid and safe for the existing CSV pipeline.
SCORESHEET_FIELDS = ["item_id", "article_title", *CRITERIA,
                     "hallucination_present", "hallucination_notes", "comment"]


# ---------------------------------------------------------------------------
# Row identity + dedupe
# ---------------------------------------------------------------------------
# In plain English: data/evaluations.jsonl can contain more than one row for
# the exact same (article, AI summarizer) pair — e.g. re-running the judges.
# The two functions below figure out "what counts as the same thing" and
# "if there are duplicates, which one do we actually use."

def _row_key(row: dict[str, Any]) -> tuple[str, str, str]:
    """Identity of the thing being reviewed: one (doi, summarizer, input_source).

    In plain English: builds a simple "fingerprint" — three text values
    stuck together as a tuple (a fixed, ordered mini-list) — that uniquely
    names one (article, which AI wrote the summary, which text version).
    Two rows with the same fingerprint are describing the same reviewable
    item.

    Matches eval_instances/reliability's join key, so a sampled evaluation row
    can be matched back to the EvaluationInstance that carries its full text.
    """
    return (
        str(row.get("doi", "")).strip(),
        str(row.get("summarizer", "")).strip(),
        str(row.get("input_source") or (row.get("strata") or {}).get("input_source") or "processed"),
    )


def _dedupe_rows(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep one evaluation row per (doi, summarizer, input_source): the latest.

    A multi-judge jury run or a --no-resume rerun can leave more than one row
    for the same item; the reviewer should see one summary per item, so we
    take the most recently written row (ISO-8601 timestamps sort correctly as
    plain strings).
    """
    # latest: a dictionary mapping each fingerprint (from _row_key) to the
    # single newest row seen so far for it. We walk every row once; if we've
    # never seen this fingerprint, or this row is newer than what we stored,
    # it replaces the stored one.
    latest: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = _row_key(row)
        if not key[0] or not key[1]:
            continue
        existing = latest.get(key)
        if existing is None or str(row.get("timestamp") or "") >= str(existing.get("timestamp") or ""):
            latest[key] = row
    return list(latest.values())


# ---------------------------------------------------------------------------
# Stratified sampler
# ---------------------------------------------------------------------------
# In plain English: "stratified sampling" means splitting the pool of
# articles into subgroups (strata) — e.g. by journal or species — and making
# sure the sample pulls from every subgroup instead of randomly landing all
# in one subgroup by chance. The functions below implement that, plus
# priority for any row already flagged as needing human eyes on it.

def _strata_group_key(row: dict[str, Any]) -> tuple[str, ...]:
    """A printable composite key across STRATIFICATION_FIELDS for grouping.

    In plain English: reads this row's journal/species/study-design/etc.
    values (STRATIFICATION_FIELDS, defined in eval_instances.py) and glues
    them into one tuple, so two rows sharing every one of those values get
    the exact same tuple back — that tuple is their "group" for sampling.
    """
    strata = row.get("strata") or {}

    def norm(value: Any) -> str:
        # A field's stored value might be a single string (e.g. "JVIM") or a
        # list (e.g. ["Canine", "Feline"] for species). This turns either
        # shape into one plain, comparable string.
        if isinstance(value, list):
            return ", ".join(str(v) for v in value) if value else "unknown"
        return str(value or "unknown")

    return tuple(norm(strata.get(field, "unknown")) for field in STRATIFICATION_FIELDS)


def _select_units_stratified(
    units: list[Any], *, sample_size: int, rng: random.Random,
    flagged_fn: Any, strata_key_fn: Any, sort_key_fn: Any,
) -> list[Any]:
    """Flagged-first, then stratified round-robin selection over any unit type.

    Shared by ``sample_rows_for_review``'s two ``sample_unit`` modes: a
    "unit" is one row in "items" mode, or one article's whole list of rows in
    "articles" mode. ``flagged_fn``/``strata_key_fn``/``sort_key_fn`` let the
    caller define what "flagged" and "strata" mean for its unit type, while
    the flagged-first-then-stratified-round-robin selection algorithm itself
    (see module-level sampler docs) stays in one place.

    1. Every unit ``flagged_fn`` marks True is included first
       (deterministically shuffled and capped at ``sample_size`` if there are
       more flagged units than the requested sample).
    2. Remaining slots are filled round-robin across ``strata_key_fn``
       groups, so the sample spans subgroups instead of being dominated by
       whichever group happens to have the most units.

    In plain English: this is a generic "pick N things fairly" helper. It
    doesn't care whether a "unit" is one row or a whole article's group of
    rows — the caller hands in three small functions (``flagged_fn`` etc.)
    that answer "is this flagged?", "which group is it in?", "how do I sort
    it?", and this function does the actual picking logic once, so it's
    never duplicated for the two different sampling modes below.
    ``rng`` is a ``random.Random(seed)`` object — a random-number generator
    that, given the same starting seed, always produces the same sequence of
    "random" choices, which is what makes the whole sample reproducible.
    """
    # Step 1: split units into "flagged" (already marked as needing a human
    # look) and "remainder" (everything else). Sorting first, then shuffling
    # with the seeded rng, means the shuffle is reproducible: same input +
    # same seed always produces the same order.
    flagged = sorted((u for u in units if flagged_fn(u)), key=sort_key_fn)
    remainder = sorted((u for u in units if not flagged_fn(u)), key=sort_key_fn)
    rng.shuffle(flagged)
    rng.shuffle(remainder)

    # Step 2: take flagged units first, up to the full sample size. If there
    # are more flagged units than we have room for, only the first
    # sample_size (after shuffling) make the cut.
    selected: list[Any] = flagged[:sample_size]
    remaining_slots = sample_size - len(selected)
    if remaining_slots <= 0:
        return selected

    # Step 3: bucket the leftover (non-flagged) units by their stratum group
    # (e.g. by journal). ``groups`` is a defaultdict(list): asking for a
    # group that doesn't exist yet silently gives you a fresh empty list
    # instead of raising an error, which is convenient for "append to
    # whichever bucket this belongs in" code like the loop below.
    groups: dict[tuple[str, ...], list[Any]] = defaultdict(list)
    for unit in remainder:
        groups[strata_key_fn(unit)].append(unit)
    group_keys = sorted(groups)
    rng.shuffle(group_keys)

    # Step 4: round-robin across the groups — take one unit from each group
    # in turn, looping back around, until either the remaining slots are
    # filled or every group has been fully drained. This is what guarantees
    # the sample doesn't get dominated by whichever group happens to have
    # the most rows.
    pool: list[Any] = []
    idx = 0
    progressed = True
    while progressed and len(pool) < remaining_slots:
        progressed = False
        for group_key in group_keys:
            bucket = groups[group_key]
            if idx < len(bucket):
                pool.append(bucket[idx])
                progressed = True
                if len(pool) >= remaining_slots:
                    break
        idx += 1

    selected.extend(pool)
    return selected


def _interleave_article_groups(
    groups: list[list[dict[str, Any]]], *, rng: random.Random,
) -> list[dict[str, Any]]:
    """Round-robin interleave each selected article's rows across the sample.

    Used only by ``sample_unit="articles"``: without this, all of one
    article's provider rows would sit consecutively in the sampled order (and
    therefore consecutively in the rendered packet), which is exactly the
    back-to-back comparison the blind protocol avoids (see "Why sampling can
    reuse the same article across multiple items" in
    docs/phase5/human_validation.md). Shuffling the article order, then
    taking one row per article per round, guarantees two rows sharing a
    ``doi`` are never adjacent whenever two or more articles are selected —
    with a single selected article there is nothing to interleave with, so
    its rows are necessarily consecutive.

    In plain English: imagine three articles, each with 3 AI summaries to
    score (9 rows total). Without this step, a reviewer would read
    Article A's 3 summaries back-to-back, then B's 3, then C's 3 — making it
    easy to compare "AI #1's summary of A" against "AI #2's summary of A"
    right next to each other, which could tip off a reviewer or bias their
    score. This function shuffles the article order and then deals out one
    row from each article per "round" (like dealing playing cards one at a
    time to each player), so the same article's summaries end up spread
    apart in the final packet.
    """
    order = list(range(len(groups)))
    rng.shuffle(order)
    shuffled_groups = [groups[i] for i in order]

    interleaved: list[dict[str, Any]] = []
    idx = 0
    progressed = True
    while progressed:
        progressed = False
        for group in shuffled_groups:
            if idx < len(group):
                interleaved.append(group[idx])
                progressed = True
        idx += 1
    return interleaved


def sample_rows_for_review(
    rows: Iterable[dict[str, Any]], *, sample_size: int, seed: int,
    sample_unit: str = "items",
) -> list[dict[str, Any]]:
    """Stratified sample of evaluation rows, preferring flagged-for-review rows.

    In plain English: this is the function that decides WHICH already-judged
    articles/summaries get sent out for human review, out of everything in
    data/evaluations.jsonl. It always prefers anything already flagged as
    needing a closer look, then fills the rest of the sample so every
    journal/species/etc. subgroup gets fair representation.

    ``sample_unit`` controls what ``sample_size`` counts:

    - ``"items"`` (default) — one row (one article + one provider's summary)
      per count. Every deduped row with ``requires_human_review=True`` is
      included first; remaining slots are filled round-robin across the
      existing strata groups (species / study_design / clinical_topic /
      journal / input_source, see ``eval_instances.STRATIFICATION_FIELDS``),
      so the sample spans subgroups instead of being dominated by whichever
      journal/strata combination happens to have the most rows.
    - ``"articles"`` — one distinct article (``doi``) per count, but EVERY
      provider row found for that article is included, so a reviewer reads
      each sampled article once and scores every provider's summary of it
      (e.g. 5 articles x 3 providers = 15 scored items from 5 articles
      actually read). An article is "flagged" if any of its rows is; the
      same stratified round-robin selection runs at article granularity.
      Sibling rows (same ``doi``) are round-robin interleaved with other
      selected articles' rows in the returned order so they are never
      adjacent in the rendered packet, preserving independent, non-
      comparative scoring — see "Why sampling can reuse the same article
      across multiple items" in docs/phase5/human_validation.md.

    Deterministic for a given ``seed`` so re-running the export with the same
    seed reproduces the same sample (useful when a reviewer sheet is lost and
    needs regenerating).
    """
    if sample_unit not in ("items", "articles"):
        raise ValueError(f"sample_unit must be 'items' or 'articles', got {sample_unit!r}")

    # First, collapse any duplicate rows for the same item down to one (see
    # _dedupe_rows above). If there's nothing to sample from, or the caller
    # asked for zero items, there's nothing more to do.
    deduped = _dedupe_rows(rows)
    if sample_size <= 0 or not deduped:
        return []

    # random.Random(seed) creates a random-number generator seeded with a
    # fixed starting number. Unlike plain ``random``, the same seed always
    # produces the exact same sequence of "random" picks — that's what makes
    # sample_rows_for_review's output reproducible run to run.
    rng = random.Random(seed)

    if sample_unit == "items":
        # "items" mode: each row (one article + one AI's summary of it) is
        # its own unit, sampled directly.
        return _select_units_stratified(
            deduped, sample_size=sample_size, rng=rng,
            flagged_fn=lambda r: bool(r.get("requires_human_review")),
            strata_key_fn=_strata_group_key,
            sort_key_fn=_row_key,
        )

    # sample_unit == "articles": group all rows by their article (doi) first,
    # so an "article" — with every AI's summary of it bundled together — is
    # what gets sampled, not an individual row.
    article_groups = _group_rows_by_article(deduped)

    # Pick which article-groups make the sample, then spread each selected
    # article's several summary-rows apart in the final order (see
    # _interleave_article_groups) so they aren't read back-to-back.
    selected_groups = _select_units_stratified(
        article_groups, sample_size=sample_size, rng=rng,
        flagged_fn=lambda g: any(r.get("requires_human_review") for r in g),
        strata_key_fn=lambda g: _strata_group_key(g[0]),
        sort_key_fn=lambda g: _row_key(g[0]),
    )
    return _interleave_article_groups(selected_groups, rng=rng)


def _group_rows_by_article(rows: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    """Group already-deduped rows into one list per article (``doi``).

    In plain English: takes a flat list of evaluation rows and bundles every
    row that shares the same article (``doi``) into its own sub-list — e.g.
    three providers' summaries of the same paper become one three-row group.
    Each group's rows are sorted by ``_row_key`` for a deterministic order.
    Shared by ``sample_rows_for_review``'s ``"articles"`` mode and the
    per-journal quota sampler below, so the grouping logic lives in one place.
    """
    by_doi: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        doi = str(row.get("doi", "")).strip()
        if doi:
            by_doi[doi].append(row)
    return [sorted(rows_, key=_row_key) for rows_ in by_doi.values()]


def _article_group_journal(group: list[dict[str, Any]]) -> str:
    """The journal an article-group belongs to, read from its lead row.

    In plain English: every row in a group is the same article, so they all
    carry the same ``strata.journal`` value — this just reads it off the
    first row.
    """
    return str((group[0].get("strata") or {}).get("journal", "unknown"))


def sample_articles_from_journal_quota(
    rows: Iterable[dict[str, Any]], *, quota: dict[str, int], seed: int,
    on_shortfall: str = "error",
    backfill_pool: Iterable[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Sample articles to an EXACT per-journal count, given explicitly.

    In plain English: unlike ``sample_rows_for_review``'s composite-strata
    round robin (which does not guarantee an exact split), this function is
    handed a precise target per journal (e.g. ``{"JVIM": 2, "JAVMA": 2, ...}``)
    and either fills every journal to that exact count or refuses to guess:

    - ``on_shortfall="error"`` (the real study export's mode): if any
      journal's evaluated-article pool can't supply its quota, nothing is
      returned — :class:`JournalQuotaError` is raised naming every short
      journal at once, so the "equal per journal" promise is never silently
      broken.
    - ``on_shortfall="backfill"`` (the pilot dry-run's mode): take whatever
      each short journal has, then fill the remaining shortfall from the
      rest of ``rows`` and, if that's still not enough, from
      ``backfill_pool`` too (possibly reusing an already-sampled article as
      a last resort) — printing a loud warning either way. Appropriate for a
      rehearsal tool whose whole point is a still-growing dev pool.
      ``backfill_pool`` lets a caller offer a wider fallback (e.g. the full
      eligible pool, including articles already shown to a previous tester)
      without those articles being eligible for the *primary*, non-backfill
      selection above — it is only ever consulted once ``rows`` itself has
      been exhausted. Defaults to ``rows`` itself when omitted.

    A journal with ``quota[j] == 0`` is left alone entirely (nothing to pick,
    nothing that can be short). A journal not present in ``quota`` at all is
    ignored the same way. Every selected article's summaries are flagged-first
    then seed-shuffled within their own journal (mirrors
    ``sample_rows_for_review``'s selection algorithm via
    ``_select_units_stratified``), and the final combined selection is
    interleaved (``_interleave_article_groups``) so two summaries of the same
    article are never adjacent in the rendered packet.
    """
    if on_shortfall not in ("error", "backfill"):
        raise ValueError(f"on_shortfall must be 'error' or 'backfill', got {on_shortfall!r}")

    deduped = _dedupe_rows(rows)
    article_groups = _group_rows_by_article(deduped)
    by_journal: dict[str, list[list[dict[str, Any]]]] = defaultdict(list)
    for group in article_groups:
        by_journal[_article_group_journal(group)].append(group)

    rng = random.Random(seed)
    selected_groups: list[list[dict[str, Any]]] = []
    shortfall: dict[str, tuple[int, int]] = {}

    for journal in sorted(quota):
        needed = quota[journal]
        if needed <= 0:
            continue
        available_groups = by_journal.get(journal, [])
        if len(available_groups) < needed:
            shortfall[journal] = (len(available_groups), needed)
            if on_shortfall == "backfill":
                # Take everything this journal has; the rest of the
                # shortfall is made up from the general pool below.
                selected_groups.extend(
                    _select_units_stratified(
                        available_groups, sample_size=len(available_groups), rng=rng,
                        flagged_fn=lambda g: any(r.get("requires_human_review") for r in g),
                        strata_key_fn=lambda g: ("journal",),
                        sort_key_fn=lambda g: _row_key(g[0]),
                    )
                )
            continue
        selected_groups.extend(
            _select_units_stratified(
                available_groups, sample_size=needed, rng=rng,
                flagged_fn=lambda g: any(r.get("requires_human_review") for r in g),
                strata_key_fn=lambda g: ("journal",),
                sort_key_fn=lambda g: _row_key(g[0]),
            )
        )

    if shortfall and on_shortfall == "error":
        raise JournalQuotaError(shortfall)

    if shortfall and on_shortfall == "backfill":
        total_missing = sum(needed - available for available, needed in shortfall.values())
        already_chosen_dois = {g[0].get("doi") for g in selected_groups}
        leftover_groups = [
            g for g in article_groups if g[0].get("doi") not in already_chosen_dois
        ]
        # Still short after exhausting `rows` itself? Fall back to the wider
        # `backfill_pool` (e.g. articles a previous tester already saw) as
        # the last resort, rather than silently under-filling.
        if len(leftover_groups) < total_missing and backfill_pool is not None:
            wider_groups = _group_rows_by_article(_dedupe_rows(backfill_pool))
            already_or_leftover_dois = already_chosen_dois | {
                g[0].get("doi") for g in leftover_groups
            }
            leftover_groups += [
                g for g in wider_groups if g[0].get("doi") not in already_or_leftover_dois
            ]
        backfill_groups = _select_units_stratified(
            leftover_groups, sample_size=total_missing, rng=rng,
            flagged_fn=lambda g: any(r.get("requires_human_review") for r in g),
            strata_key_fn=lambda g: ("any",),
            sort_key_fn=lambda g: _row_key(g[0]),
        )
        if backfill_groups:
            print(
                f"[human_review] WARNING: the evaluated pool could not supply the exact "
                f"per-journal quota; backfilled {len(backfill_groups)} article(s) from "
                "other journals: "
                + ", ".join(
                    f"{journal} short by {needed - available}"
                    for journal, (available, needed) in sorted(shortfall.items())
                )
                + ". Run more 'evaluate' rounds to widen the thin journal(s)."
            )
            selected_groups.extend(backfill_groups)

    return _interleave_article_groups(selected_groups, rng=rng)


def sample_articles_equal_per_journal(
    rows: Iterable[dict[str, Any]], *, sample_size: int, seed: int,
    journals: tuple[str, ...] = STUDY_JOURNALS, on_shortfall: str = "error",
) -> list[dict[str, Any]]:
    """Sample ``sample_size`` articles split EXACTLY evenly across ``journals``.

    In plain English: the simple, common-case entry point — "give me N
    articles, N/len(journals) from each journal." ``sample_size`` must be a
    multiple of ``len(journals)`` (5, 10, 25 for the study's 5 journals) or
    it can't split evenly; anything else raises ``ValueError`` pointing at
    :data:`ARTICLE_COUNT_CHOICES`. Internally this is a thin wrapper around
    :func:`sample_articles_from_journal_quota` with an equal quota dict built
    from ``sample_size``.
    """
    if sample_size <= 0 or sample_size % len(journals) != 0:
        raise ValueError(
            f"sample_size must be a positive multiple of {len(journals)} (one of "
            f"{ARTICLE_COUNT_CHOICES}) to split evenly across {len(journals)} journals, "
            f"got {sample_size}."
        )
    per_journal = sample_size // len(journals)
    quota = {journal: per_journal for journal in journals}
    return sample_articles_from_journal_quota(
        rows, quota=quota, seed=seed, on_shortfall=on_shortfall,
    )


# ---------------------------------------------------------------------------
# Join sampled rows to their source text
# ---------------------------------------------------------------------------
# In plain English: sample_rows_for_review() above only picked WHICH
# (article, AI) pairs to review — those rows are still just short summary
# records from evaluations.jsonl, without the full article text or the full
# candidate summary text attached. The code below fetches that full text and
# bundles everything a reviewer needs into one tidy object per item.

@dataclass(frozen=True)
class ReviewItem:
    """One blind review unit: an anonymized id plus everything a reviewer needs.

    ``summarizer``/``judge``/LLM score fields are carried on this object only
    so ``build_unblinding_key()`` can record them — the rendering functions
    below (``render_packet_markdown``, ``render_scoresheet_csv``) never read
    those fields, which is the structural guarantee of the blind protocol.

    In plain English: a dataclass is like a labeled form with fixed fields —
    you declare the field names once (below) and Python builds the rest
    (constructor, printing, etc.) automatically. ``frozen=True`` means once a
    ReviewItem is created, its fields can't be changed afterwards — like
    filling out a form in permanent ink. This particular "form" bundles
    together everything there is to know about one reviewable
    (article, AI-written summary) pair, including some fields (summarizer,
    judge, LLM scores) that are ONLY used to build the private unblinding
    key — never shown to a reviewer.
    """

    item_id: str
    doi: str
    # The article's human-readable title. Shown to the reviewer (packet heading
    # + a reference column in the scoresheet) so they can tell items apart, and
    # blind-safe: a title identifies the *article*, never which AI wrote the
    # summary — the only thing the blind protocol hides. Defaults to "" so the
    # renderers degrade gracefully when a summary row carried no title.
    title: str
    summarizer: str
    judge: str
    rubric_version: str
    input_source: str
    reference_text: str
    candidate_summary: str
    strata: dict[str, Any]
    requires_human_review: bool
    llm_jury_score: float | None
    llm_jury_score_weighted: float | None
    llm_jury_score_unweighted: float | None
    llm_criteria_scores: dict[str, Any]


def _build_instance_lookup(
    summaries_path: Path,
    summaries_txt_dir: Path,
    dev_tests_summaries_txt_dir: Path,
) -> dict[tuple[str, str, str], EvaluationInstance]:
    """Combine both text sources evaluate() can score from into one lookup.

    Mirrors run_phase3.py's EVAL_INPUT_MODE duality (jsonl vs. summarize-all's
    .txt folders): a sampled evaluations.jsonl row may have come from either
    path, so both are tried, jsonl-sourced instances taking precedence
    (``setdefault``) since that is the original, still-default pipeline.

    In plain English: builds one big "phone book" dictionary you can look up
    by (article, AI, text-version) to instantly get that pairing's full
    article text and full AI summary text, regardless of which of the two
    places on disk that text actually lives in. ``dict.setdefault(key, val)``
    only inserts ``val`` if ``key`` isn't already in the dictionary — so if
    the same (article, AI, version) combination shows up in both places, the
    first (jsonl) source wins and the second is ignored.
    """
    lookup: dict[tuple[str, str, str], EvaluationInstance] = {}
    for instance in iter_evaluation_instances(summaries_path=summaries_path):
        lookup.setdefault((instance.doi, instance.summarizer, instance.input_source), instance)
    for folder in (summaries_txt_dir, dev_tests_summaries_txt_dir):
        for instance in iter_summarize_all_instances(folder):
            lookup.setdefault((instance.doi, instance.summarizer, instance.input_source), instance)
    return lookup


def _build_review_items(
    sampled_rows: list[dict[str, Any]],
    lookup: dict[tuple[str, str, str], EvaluationInstance],
) -> tuple[list[ReviewItem], list[dict[str, Any]]]:
    """Join sampled evaluation rows to their reference/candidate text.

    Returns (items, skipped_rows). A row is skipped when its
    (doi, summarizer, input_source) can't be matched to a live instance —
    e.g. data/summaries.jsonl or the summarize-all .txt folders have since
    been moved, regenerated, or pruned. Skipping (rather than raising) keeps
    export usable even when the corpus has drifted since evaluation; the
    caller prints a warning with the count so nothing is silently lost.

    In plain English: for each sampled row, look up its full text in the
    lookup table built above, and package it into a ReviewItem with a
    friendly, anonymous id like "item_001". ``enumerate(sampled_rows, start=1)``
    just numbers the rows starting at 1 instead of 0, so the first item is
    "item_001" not "item_000".
    """
    items: list[ReviewItem] = []
    skipped: list[dict[str, Any]] = []
    for n, row in enumerate(sampled_rows, start=1):
        key = _row_key(row)
        instance = lookup.get(key)
        if instance is None:
            skipped.append(row)
            continue
        items.append(ReviewItem(
            item_id=f"item_{n:03d}",
            doi=key[0],
            title=str(instance.summary_record.get("title", "") or ""),
            summarizer=key[1],
            judge=str(row.get("judge", "")),
            rubric_version=str(row.get("rubric_version", "")),
            input_source=key[2],
            reference_text=instance.reference_text,
            candidate_summary=instance.candidate_summary,
            strata=row.get("strata") or {},
            requires_human_review=bool(row.get("requires_human_review")),
            llm_jury_score=row.get("jury_score"),
            llm_jury_score_weighted=row.get("jury_score_weighted"),
            llm_jury_score_unweighted=row.get("jury_score_unweighted"),
            llm_criteria_scores=row.get("criteria_scores") or {},
        ))
    return items, skipped


# ---------------------------------------------------------------------------
# Blind rendering (packet + scoresheet) — no summariser/judge identity
# ---------------------------------------------------------------------------
# In plain English: the functions in this section turn ReviewItem objects
# into the actual text files a reviewer opens and reads (a Markdown "packet"
# document, and a fillable CSV/Excel scoresheet). "Markdown" is a simple
# plain-text format (using things like "# Heading" and "**bold**") that
# renders nicely when opened in many text editors or on GitHub. Every
# function here is written to only ever read the handful of ReviewItem
# fields that are safe to show a reviewer (id, article title, article text,
# candidate summary) — never which AI wrote the summary.

def render_packet_markdown(
    items: list[ReviewItem], *, reviewer_id: int,
    scoresheet_filename: str | None = None,
) -> str:
    """Render the blind reading packet: item_id + article title + text + summary.

    In plain English: builds the full text of the old-style, single-file
    reading packet (one long document listing every item's article text and
    candidate summary) as one big string, ready to be saved to disk. (The
    export flow used today instead calls ``render_packet_index_markdown`` +
    ``write_item_folders``, which split this into one folder per item — this
    function is kept for anything that still wants one flat file.)

    Deliberately takes ``ReviewItem`` objects but only ever reads
    ``item_id``, ``title``, ``reference_text``, and ``candidate_summary`` —
    never the summariser/judge identity, the same structural guarantee as
    ``evaluator.build_judge_prompt()`` not accepting a summariser parameter at
    all. (``title`` names the *article*, not the model, so it is blind-safe.)

    ``scoresheet_filename`` is the exact filename the "score each item in ..."
    instruction points reviewers at. It defaults to the real export's
    ``scoresheet_reviewer_{reviewer_id}.csv``; the pilot export passes its own
    ``scoresheet_human{N}.xlsx`` so the packet never points at a file that
    isn't in the folder.
    """
    scoresheet_name = scoresheet_filename or f"scoresheet_reviewer_{reviewer_id}.csv"
    lines = [
        f"# Human Validation Packet — Reviewer {reviewer_id}",
        "",
        "**For the full guide** — what each column means, a worked example, and "
        "why everything here is blind — see `REVIEWER_GUIDE.md` in this same "
        "folder. Read it once before you start. What follows is just a quick "
        "reference for while you're scoring.",
        "",
        "This packet is BLIND: each item shows only the original article text "
        "and one candidate summary. You are not told which system (or person) "
        "wrote the summary — please do not try to guess, and score each item "
        "independently of the others.",
        "",
        "**Some articles may appear more than once**, each time paired with a "
        "different summary. Score every occurrence completely independently, "
        "as if it were the first time you'd read that article — don't look "
        "back at an earlier score for the same article.",
        "",
        f"Score each item in `{scoresheet_name}`. Use the "
        "`item_id` heading below to find the matching row in that file.",
        "",
        "Scoring scale (1-5, higher is better) for each of the five columns "
        "in the scoresheet:",
        "",
        "- **faithfulness** — does the summary avoid claims not supported by the article?",
        "- **completeness** — does it cover the article's key findings?",
        "- **clinical_usefulness** — would this help a veterinary clinician or researcher?",
        "- **clarity** — is it well-organized and easy to read?",
        "- **safety** — could a vet reader be clinically misled by anything stated or omitted?",
        "",
        "Also mark `hallucination_present` (yes/no) if the summary states something "
        "the article does not support, with a short note, and use `comment` for "
        "anything else worth flagging.",
        "",
        "---",
        "",
    ]
    for item in items:
        lines.append(f"## {item.item_id}")
        lines.append("")
        # Blind-safe: the article title identifies the paper, not the model.
        # Skipped entirely when absent so the packet degrades gracefully.
        if item.title.strip():
            lines.append(f"**Article:** {item.title.strip()}")
            lines.append("")
        lines.append("### Original article text")
        lines.append("")
        lines.append(item.reference_text.strip())
        lines.append("")
        lines.append("### Candidate summary")
        lines.append("")
        lines.append(item.candidate_summary.strip())
        lines.append("")
        lines.append("---")
        lines.append("")
    return "\n".join(lines) + "\n"


def render_reviewer_guide_markdown() -> str:
    """Read the standalone, zero-jargon reviewer guide (docs/booklet ch. 7).

    In plain English: this function doesn't build any text itself — it just
    opens an existing instructions file from disk and returns its contents
    unchanged, so every reviewer folder gets an identical copy of the guide.

    Export copies this verbatim into every humanN/ folder as
    REVIEWER_GUIDE.md, so a reviewer handed only their own folder still gets
    the full guide — not just packet.md's compact quick-reference bullets.
    The doc file itself is the single source of truth; this function only
    distributes it, so a future edit to the guide propagates on the next
    export with no code change needed here.
    """
    if not REVIEWER_GUIDE_PATH.exists():
        raise FileNotFoundError(
            f"Reviewer guide not found at {REVIEWER_GUIDE_PATH}. Export cannot "
            "produce a self-contained reviewer folder without it."
        )
    return REVIEWER_GUIDE_PATH.read_text(encoding="utf-8")


def render_scoresheet_csv(items: list[ReviewItem]) -> str:
    """Render the fillable CSV scoresheet: one row per item_id.

    ``item_id`` and ``article_title`` are pre-filled reference cells; the
    scored columns are left blank for the reviewer to fill.

    In plain English: builds a plain-text spreadsheet (CSV = "comma
    separated values", something like Excel or Google Sheets can open) with
    one row per reviewable item, ready for a reviewer to type their 1-5
    scores into. ``io.StringIO()`` is a way to build up text in memory as if
    it were being written to a file, without actually touching the disk —
    handy when you want the finished text back as a string.
    ``csv.DictWriter`` is a standard-library tool that writes one dictionary
    per row into CSV format, matching each dictionary's keys to the right
    column, so you don't have to worry about comma-escaping or column order
    yourself.
    """
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=SCORESHEET_FIELDS)
    writer.writeheader()
    for item in items:
        row = {field: "" for field in SCORESHEET_FIELDS}
        row["item_id"] = item.item_id
        row["article_title"] = item.title
        writer.writerow(row)
    return buf.getvalue()


# In plain English: how wide (in Excel's character-count units) each column
# should be drawn when the .xlsx scoresheet is opened, so long text (title,
# notes, comment) doesn't look squashed and short 1-5 score columns don't
# waste space. Purely cosmetic — doesn't affect what data is stored.
_XLSX_COLUMN_WIDTHS = {
    "item_id": 12,
    "article_title": 55,
    "faithfulness": 13,
    "completeness": 13,
    "clinical_usefulness": 18,
    "clarity": 13,
    "safety": 13,
    "hallucination_present": 20,
    "hallucination_notes": 42,
    "comment": 42,
}


def render_scoresheet_xlsx(
    items: list[ReviewItem], path: Path, versions: dict[str, tuple[int, int]] | None = None,
) -> None:
    """Write the fillable .xlsx scoresheet — same columns as the CSV, easier to fill.

    Shares ``SCORESHEET_FIELDS`` with :func:`render_scoresheet_csv` so the two
    renderers can never drift on which columns exist or their order. On top of
    the plain CSV it adds reviewer-usability affordances a flat file can't:

    - Dropdown validation: 1-5 on the five criteria columns, yes/no on
      ``hallucination_present`` — a reviewer picks instead of typing, so a
      stray ``44`` or ``good`` can't get entered in the first place (the CSV
      pipeline only catches those *after the fact*, in ingest).
    - A frozen, bold header row and per-column widths with wrapped text, so the
      long article title / notes / comment cells stay readable.

    ``item_id`` and ``article_title`` are pre-filled reference cells; the scored
    columns are left blank. Written straight to ``path`` (xlsx is binary, unlike
    the CSV renderer's string return). ``openpyxl`` is imported lazily so the
    rest of this module — and the real CSV export/ingest — never depend on it.

    ``versions`` (item_id -> (k, n) from :func:`summary_versions`) disambiguates
    the rows: when an article contributes ``n > 1`` summaries, its
    ``article_title`` cell is suffixed ``— summary version k of n`` so the three
    same-title rows are distinguishable. The suffix is a version *number*, never
    a provider name, so it stays blind. ``item_id`` is untouched (ingest keys on
    it), so the label is purely a reviewer-usability aid.

    In plain English: this builds an actual Excel (.xlsx) file, cell by cell,
    using the third-party ``openpyxl`` library (not part of core Python, so
    it needs to be installed separately — that's what the try/except below
    checks for). Think of it as programmatically doing what you'd do by hand
    in Excel: type the column headers, fill in a couple of pre-known cells
    per row, set column widths, freeze the header row so it stays visible
    while scrolling, and add dropdown menus so reviewers pick a valid score
    instead of typing something invalid.
    """
    # openpyxl is only imported here (not at the top of the file) so that
    # code which never touches .xlsx files doesn't require it to be
    # installed at all. If it's missing, turn the unhelpful ImportError into
    # a clear instruction for the user.
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:  # pragma: no cover - exercised only without openpyxl
        raise RuntimeError(
            "render_scoresheet_xlsx needs the 'openpyxl' package "
            "(pip install openpyxl). The plain-CSV scoresheet renderer has no "
            "such dependency if you'd rather avoid it."
        ) from exc

    # Create a new, empty workbook (an Excel file) with one worksheet (one
    # tab/page inside it), and name that worksheet "scoresheet".
    wb = Workbook()
    ws = wb.active
    ws.title = "scoresheet"

    header_font = Font(bold=True)
    header_align = Alignment(vertical="top", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)

    # Header row: write each column's name (SCORESHEET_FIELDS) into row 1,
    # bold it, and set that column's on-screen width.
    for col_idx, field in enumerate(SCORESHEET_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            _XLSX_COLUMN_WIDTHS.get(field, 16)
        )

    # One data row per item (starting at row 2, since row 1 is the header);
    # only the two reference cells (item_id, article_title) are pre-filled —
    # everything else is left as an empty cell for the reviewer to type into.
    for row_idx, item in enumerate(items, start=2):
        for col_idx, field in enumerate(SCORESHEET_FIELDS, start=1):
            value = None
            if field == "item_id":
                value = item.item_id
            elif field == "article_title":
                value = _article_title_with_version(item, versions)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align

    # Freeze the header row (row 1) so it stays visible at the top of the
    # screen when a reviewer scrolls down through many items.
    ws.freeze_panes = "A2"

    # Dropdown validation over the data range (row 2 .. last item row): adds
    # a clickable dropdown list to each score cell (limiting entries to
    # 1-5) and each hallucination_present cell (limiting to yes/no), so a
    # reviewer can only enter a valid value in the first place.
    last_row = len(items) + 1
    if last_row >= 2:
        field_to_letter = {
            field: get_column_letter(idx)
            for idx, field in enumerate(SCORESHEET_FIELDS, start=1)
        }
        score_dv = DataValidation(
            type="list", formula1='"1,2,3,4,5"', allow_blank=True,
            errorTitle="Invalid score", error="Pick a whole number from 1 to 5.",
            promptTitle="Score", prompt="1 (poor) to 5 (excellent).",
        )
        yesno_dv = DataValidation(
            type="list", formula1='"yes,no"', allow_blank=True,
            errorTitle="Invalid entry", error="Pick 'yes' or 'no'.",
            promptTitle="Hallucination present?",
            prompt="Does the summary state something the article doesn't support?",
        )
        ws.add_data_validation(score_dv)
        ws.add_data_validation(yesno_dv)
        for criterion in CRITERIA:
            letter = field_to_letter[criterion]
            score_dv.add(f"{letter}2:{letter}{last_row}")
        hp_letter = field_to_letter["hallucination_present"]
        yesno_dv.add(f"{hp_letter}2:{hp_letter}{last_row}")

    # Make sure the destination folder exists, then save the finished
    # workbook to disk as a real .xlsx file at ``path``.
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------------------
# Per-item folders + navigation index (nested reviewer-folder layout)
# ---------------------------------------------------------------------------
# In plain English: this section builds the actual folder structure a
# reviewer receives — one sub-folder per reviewable item (containing the
# article text and the candidate summary as separate files), a "packet.md"
# index page that links to all of them, and copies of the original PDFs.

def summary_versions(items: list[ReviewItem]) -> dict[str, tuple[int, int]]:
    """Map each ``item_id`` to ``(version, total)`` within its article.

    A reviewer folder holds several summaries of the same article (one per
    provider), so three rows can share an ``article_title``. This assigns each a
    blind-safe 1-based *version* — the k-th time that ``doi`` appears in the
    already-ordered ``items`` list — plus the article's total summary count, so
    the packet index and scoresheet can label them "version k of n". Because the
    export orders items provider-major (see ``_interleave_article_groups``), an
    article's versions come out in a stable order; the number never names a
    provider, so the blind protocol holds.

    In plain English: if the same article was summarized by 3 different AIs,
    a reviewer will see that article 3 times, each with a different summary.
    To avoid confusing them, each occurrence gets labeled "version 1 of 3",
    "version 2 of 3", "version 3 of 3" — never which AI made which version.
    This function works out those numbers by counting how many times each
    article's doi appears in the list, in order.
    """
    # totals: how many summaries exist in total for each article (counted
    # first, in one pass over every item).
    totals: dict[str, int] = defaultdict(int)
    for item in items:
        totals[item.doi] += 1
    # seen: a running counter of how many of THIS article's items we've
    # passed so far, used to assign "version 1", "version 2", etc. in order
    # as we walk through the list a second time.
    seen: dict[str, int] = defaultdict(int)
    versions: dict[str, tuple[int, int]] = {}
    for item in items:
        seen[item.doi] += 1
        versions[item.item_id] = (seen[item.doi], totals[item.doi])
    return versions


def _article_title_with_version(
    item: ReviewItem, versions: dict[str, tuple[int, int]] | None,
) -> str:
    """The reviewer-facing article label: title, plus a version suffix if repeated.

    In plain English: returns the text a reviewer actually sees for an
    article's title cell — just the plain title normally, or the title with
    " — summary version k of n" tacked on when the article appears more than
    once in this reviewer's packet.
    """
    title = item.title
    if versions:
        k, n = versions.get(item.item_id, (1, 1))
        if n > 1:
            suffix = f"summary version {k} of {n}"
            title = f"{title} — {suffix}" if title.strip() else suffix
    return title


def write_item_folders(
    items: list[ReviewItem], base_dir: Path,
    versions: dict[str, tuple[int, int]] | None = None,
) -> list[Path]:
    """Write one ``item_id/`` folder per review item: ``article.md`` + ``summary.md``.

    This is the nested reviewer layout: instead of one long ``packet.md`` with
    every article+summary inline, each item gets its own folder holding the
    full original article text (``article.md``) and the single candidate summary
    to score against it (``summary.md``). The folder is named by ``item_id`` so
    it lines up exactly with the scoresheet's ``item_id`` column.

    Blind-safe by construction: like ``render_packet_markdown`` it reads only
    ``item_id``, ``title``, ``reference_text``, ``candidate_summary`` and the
    version number — never the summariser/judge identity.

    In plain English: for every review item, this creates a folder named
    after its item_id (e.g. ``item_003/``) containing two files:
    ``article.md`` (the full original article text) and ``summary.md`` (the
    one AI-written summary to score against it). This is what a reviewer
    actually opens and reads.
    """
    if versions is None:
        versions = summary_versions(items)
    base_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for item in items:
        folder = base_dir / item.item_id
        folder.mkdir(parents=True, exist_ok=True)

        # Build article.md: the article's title (if any) followed by the
        # full original article text.
        title = item.title.strip()
        article_lines: list[str] = []
        if title:
            article_lines.append(f"# {title}")
            article_lines.append("")
        article_lines.append("## Original article text")
        article_lines.append("")
        article_lines.append(item.reference_text.strip())
        article_lines.append("")
        (folder / "article.md").write_text("\n".join(article_lines) + "\n", encoding="utf-8")

        # Build summary.md: a heading (with a version label if this article
        # was summarized more than once), a short instruction, and the
        # candidate summary text itself.
        k, n = versions.get(item.item_id, (1, 1))
        heading = f"# {item.item_id} — candidate summary"
        if n > 1:
            heading += f" (version {k} of {n})"
        summary_lines = [heading, ""]
        if title:
            summary_lines.append(f"**Article:** {title}")
            summary_lines.append("")
        summary_lines.append(
            "Read `article.md` in this folder, then score THIS summary against it "
            f"by filling row `{item.item_id}` in the scoresheet. If the article "
            "appears more than once, score every version completely independently."
        )
        summary_lines.append("")
        summary_lines.append("## Candidate summary")
        summary_lines.append("")
        summary_lines.append(item.candidate_summary.strip())
        summary_lines.append("")
        (folder / "summary.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
        written.append(folder)
    return written


def render_packet_index_markdown(
    items: list[ReviewItem], *, reviewer_id: int | str,
    scoresheet_filename: str, versions: dict[str, tuple[int, int]] | None = None,
) -> str:
    """Render ``packet.md`` as a navigation index over the per-item folders.

    Replaces the old inline-everything packet: the full article text and
    summaries now live in the ``item_id/`` folders (see ``write_item_folders``),
    so this file keeps the blind intro + scoring guidance and adds a table
    pointing at each folder. Reads no summariser/judge identity — same blind
    guarantee as ``render_packet_markdown``.

    In plain English: builds ``packet.md``, the "home page" of a reviewer's
    folder — the instructions for how to score, plus a table of links (one
    row per item) to each item's own sub-folder built by
    ``write_item_folders``.
    """
    if versions is None:
        versions = summary_versions(items)
    lines = [
        f"# Human Validation Packet — Reviewer {reviewer_id}",
        "",
        "**For the full guide** — what each column means, a worked example, and "
        "why everything here is blind — see `REVIEWER_GUIDE.md` in this same "
        "folder. Read it once before you start. What follows is just a quick "
        "reference for while you're scoring.",
        "",
        "This packet is BLIND: each item is one original article paired with one "
        "candidate summary. You are not told which system (or person) wrote the "
        "summary — please do not try to guess, and score each item independently "
        "of the others.",
        "",
        "**Each item below is in its own folder** in this directory. Open the "
        "folder, read `article.md` (the original article) and `summary.md` (one "
        "candidate summary of it), then score that item's row in "
        f"`{scoresheet_filename}` — match by the `item_id`.",
        "",
        "**The same article appears more than once**, each time paired with a "
        "different candidate summary (a different *version*). Score every version "
        "completely independently, as if it were the first time you'd read that "
        "article — don't look back at an earlier score for the same article.",
        "",
        "Scoring scale (1-5, higher is better) for each of the five columns "
        "in the scoresheet:",
        "",
        "- **faithfulness** — does the summary avoid claims not supported by the article?",
        "- **completeness** — does it cover the article's key findings?",
        "- **clinical_usefulness** — would this help a veterinary clinician or researcher?",
        "- **clarity** — is it well-organized and easy to read?",
        "- **safety** — could a vet reader be clinically misled by anything stated or omitted?",
        "",
        "Also mark `hallucination_present` (yes/no) if the summary states something "
        "the article does not support, with a short note, and use `comment` for "
        "anything else worth flagging.",
        "",
        "---",
        "",
        "## Items in this packet",
        "",
        "| item_id | article | version | folder |",
        "| --- | --- | --- | --- |",
    ]
    for item in items:
        k, n = versions.get(item.item_id, (1, 1))
        version_cell = f"{k} of {n}" if n > 1 else "1 of 1"
        title = item.title.strip() or "(untitled)"
        lines.append(f"| `{item.item_id}` | {title} | {version_cell} | `{item.item_id}/` |")
    lines.append("")
    return "\n".join(lines) + "\n"


def copy_original_pdfs(
    items: list[ReviewItem],
    lookup: dict[tuple[str, str, str], EvaluationInstance],
    dest_dir: Path,
    raw_dir: Path,
) -> tuple[list[str], list[str]]:
    """Copy each item's source PDF from ``raw_dir`` into ``dest_dir``.

    Deduped by DOI (an article appears under more than one provider item).
    Returns ``(copied_names, missing_dois)``. A DOI whose PDF can't be found is
    reported and skipped, never fatal — the packet/article text still covers it;
    only the supplementary PDF is absent.

    In plain English: for every distinct article in this reviewer's packet,
    find its original PDF file (already downloaded earlier in the project,
    living in ``raw_dir``) and copy it into the reviewer's
    ``original_articles/`` folder, so a reviewer who wants to see the paper
    as originally published/formatted can. "Deduped by DOI" means if an
    article shows up 3 times (once per AI), its PDF is only copied once.
    """
    dest_dir.mkdir(parents=True, exist_ok=True)
    copied: list[str] = []
    missing: list[str] = []
    seen_dois: set[str] = set()
    for item in items:
        if item.doi in seen_dois:
            continue
        seen_dois.add(item.doi)
        instance = lookup.get((item.doi, item.summarizer, item.input_source))
        record = instance.summary_record if instance is not None else {"doi": item.doi}
        pdf_path = resolve_existing_pdf_path(raw_dir, record)
        if pdf_path is None:
            missing.append(item.doi)
            continue
        shutil.copy2(pdf_path, dest_dir / pdf_path.name)
        copied.append(pdf_path.name)
    return copied, missing


def write_review_folder(
    items: list[ReviewItem], folder: Path, *,
    reviewer_id: int | str, scoresheet_filename: str,
    versions: dict[str, tuple[int, int]] | None = None,
    raw_dir: Path | None = None,
    lookup: dict[tuple[str, str, str], EvaluationInstance] | None = None,
) -> tuple[list[str], list[str]]:
    """Write one self-contained reviewer folder in the nested per-item layout.

    Single source of truth shared by the real export and the pilot (both
    write ``humanN/`` folders) so their folder shapes can never drift. Writes:
    ``REVIEWER_GUIDE.md`` (verbatim guide), ``packet.md`` (navigation index),
    the ``.xlsx`` scoresheet (version-labeled), one ``item_id/`` folder per item
    (``article.md`` + ``summary.md``), and — when ``raw_dir`` and ``lookup`` are
    given — an ``original_articles/`` subfolder with the matched source PDFs.

    Returns ``(pdfs_copied, pdfs_missing)`` (both empty when no PDF copy runs).

    In plain English: this is the "assemble everything into one folder"
    function — it calls each of the pieces built above (guide, packet
    index, scoresheet, per-item folders, PDF copies) in the right order so
    that one call produces one complete, ready-to-hand-to-a-reviewer folder.
    """
    if versions is None:
        versions = summary_versions(items)
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "REVIEWER_GUIDE.md").write_text(
        render_reviewer_guide_markdown(), encoding="utf-8",
    )
    (folder / "packet.md").write_text(
        render_packet_index_markdown(
            items, reviewer_id=reviewer_id,
            scoresheet_filename=scoresheet_filename, versions=versions,
        ),
        encoding="utf-8",
    )
    render_scoresheet_xlsx(items, folder / scoresheet_filename, versions=versions)
    write_item_folders(items, folder, versions=versions)

    if raw_dir is not None and lookup is not None:
        return copy_original_pdfs(items, lookup, folder / "original_articles", raw_dir)
    return [], []


def build_unblinding_key(
    items: list[ReviewItem], *, seed: int, sample_size: int, reviewer_count: int,
) -> dict[str, Any]:
    """The item_id -> (doi, summarizer, judge, LLM scores) mapping.

    NEVER given to reviewers — written as a SIBLING of the ``humanN/`` folder
    it describes (never inside it), under a name
    (``unblinding_key_human{N}.json``, see ``unblinding_key_path_for``) that
    makes it obvious which folder it un-blinds, plus a loud console warning
    (see ``export_human_review``) that flags it as private. Recording
    ``llm_jury_score``/``llm_criteria_scores`` here (not anywhere a reviewer
    can see) is what lets the ingest step (``analyze_human_reviews``) compute
    human-vs-jury agreement without re-reading evaluations.jsonl.
    ``reviewer_count`` is always ``1`` here: each humanN/ folder is exactly
    one reviewer's packet (the "how many reviewers agree" question is
    answered across separate humanN keys at ingest time, not within one).

    In plain English: this builds the ONE file that remembers the secret —
    which AI actually wrote each summary a reviewer scored, plus what the AI
    judges themselves scored it. It's the answer key for the blind test.
    Because it's written next to the reviewer folder it describes (just
    under a clearly-labeled name, with a loud warning printed alongside it),
    whoever runs the export must be careful never to send this file to a
    reviewer.
    """
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "seed": seed,
        "requested_sample_size": sample_size,
        "actual_sample_size": len(items),
        "reviewer_count": reviewer_count,
        "items": {
            item.item_id: {
                "doi": item.doi,
                "summarizer": item.summarizer,
                "judge": item.judge,
                "rubric_version": item.rubric_version,
                "input_source": item.input_source,
                "strata": item.strata,
                "requires_human_review": item.requires_human_review,
                "llm_jury_score": item.llm_jury_score,
                "llm_jury_score_weighted": item.llm_jury_score_weighted,
                "llm_jury_score_unweighted": item.llm_jury_score_unweighted,
                "llm_criteria_scores": item.llm_criteria_scores,
            }
            for item in items
        },
    }


# ---------------------------------------------------------------------------
# Incremental humanN folders (shared by the real export and the pilot)
# ---------------------------------------------------------------------------
# In plain English: this section is what makes "run the export once ->
# human1/, run it again -> human2/ (never touching human1/)" work. It used to
# be pilot_human_review.py's private machinery; it now lives here so both the
# real study export and the pilot dry-run share one implementation.

_HUMAN_DIR_RE = re.compile(r"^human(\d+)$")


def next_human_number(output_dir: Path) -> int:
    """Return the next ``humanN`` number to create (max existing + 1, else 1).

    Looks inside ``output_dir`` for every already-existing ``humanN`` folder
    and returns one higher than the highest found (1 if there are none yet).
    Anything else in the directory (a key file, a stray file) is ignored.
    """
    highest = 0
    if output_dir.exists():
        for child in output_dir.iterdir():
            if child.is_dir():
                m = _HUMAN_DIR_RE.match(child.name)
                if m:
                    highest = max(highest, int(m.group(1)))
    return highest + 1


def unblinding_key_path_for(output_dir: Path, human_number: int) -> Path:
    """Path of tester N's private key: a SIBLING of ``humanN/``, never inside it."""
    return output_dir / f"unblinding_key_human{human_number}.json"


def load_previous_unblinding_key(output_dir: Path, human_number: int) -> dict[str, Any] | None:
    """Load tester ``N-1``'s private key, or None for the first tester / if absent."""
    if human_number <= 1:
        return None
    prev_path = unblinding_key_path_for(output_dir, human_number - 1)
    if not prev_path.exists():
        return None
    try:
        return json.loads(prev_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def _previous_articles_by_journal(previous_key: dict[str, Any]) -> dict[str, list[str]]:
    """Previous export's distinct articles (DOIs), grouped by journal.

    Ordered by each article's first-seen ``item_id`` (``item_001`` <
    ``item_002`` ...) so "take the first k of this journal's articles" is
    deterministic and reproducible.
    """
    items = previous_key.get("items") or {}
    by_journal: dict[str, list[str]] = defaultdict(list)
    seen_dois: set[str] = set()
    for item_id in sorted(items.keys()):
        entry = items[item_id]
        doi = str(entry.get("doi", "")).strip()
        if not doi or doi in seen_dois:
            continue
        seen_dois.add(doi)
        journal = str((entry.get("strata") or {}).get("journal", "unknown"))
        by_journal[journal].append(doi)
    return by_journal


def _previous_dois(previous_key: dict[str, Any]) -> set[str]:
    """Every distinct article (DOI) the previous tester saw, any journal."""
    return {
        str(item.get("doi", "")).strip()
        for item in (previous_key.get("items") or {}).values()
        if str(item.get("doi", "")).strip()
    }


def select_incremental_rows(
    eligible_rows: list[dict[str, Any]], *,
    human_number: int,
    sample_size: int,
    seed: int,
    overlap_ratio: float,
    previous_key: dict[str, Any] | None,
    journals: tuple[str, ...] = STUDY_JOURNALS,
    on_shortfall: str = "error",
) -> tuple[list[dict[str, Any]], int]:
    """Choose this tester's rows: a journal-proportional carry, then fill the rest.

    In plain English: for the very first tester (``human1``, no predecessor)
    this is just ``sample_articles_from_journal_quota`` with an equal quota.
    For every later tester, it carries a *fraction* of the previous tester's
    articles from EACH journal (so the carry itself stays balanced across
    journals), then asks the quota sampler for exactly the per-journal
    shortfall left over — which is why the leftover request is always a
    valid per-journal count, never a number that has to divide evenly by
    ``len(journals)`` on its own (a flat, journal-unaware carry would produce
    exactly that broken case: e.g. carry 3 of 5, ask for "2 more" split
    across 5 journals).

    Returns ``(selected_rows, overlap_units)`` where ``overlap_units`` is how
    many articles were carried over from the previous tester (0 for the
    first tester, or when ``overlap_ratio`` rounds every journal down to 0).
    """
    if sample_size <= 0 or sample_size % len(journals) != 0:
        raise ValueError(
            f"sample_size must be a positive multiple of {len(journals)} (one of "
            f"{ARTICLE_COUNT_CHOICES}), got {sample_size}."
        )
    per_journal = sample_size // len(journals)

    if human_number <= 1 or not previous_key:
        rows = sample_articles_from_journal_quota(
            eligible_rows, quota={journal: per_journal for journal in journals},
            seed=seed, on_shortfall=on_shortfall,
        )
        return rows, 0

    deduped = _dedupe_rows(eligible_rows)
    article_groups = _group_rows_by_article(deduped)
    groups_by_doi = {group[0].get("doi"): group for group in article_groups}

    prev_by_journal = _previous_articles_by_journal(previous_key)
    prev_dois = _previous_dois(previous_key)

    carried_rows: list[dict[str, Any]] = []
    carried_count: dict[str, int] = {}
    for journal in journals:
        # Only articles still present in the current eligible pool can be
        # carried — one may have dropped out of evaluations.jsonl since.
        still_eligible = [d for d in prev_by_journal.get(journal, []) if d in groups_by_doi]
        k = min(round(per_journal * overlap_ratio), len(still_eligible), per_journal)
        chosen = still_eligible[:k]
        carried_count[journal] = len(chosen)
        for doi in chosen:
            carried_rows.extend(groups_by_doi[doi])

    residual_quota = {journal: per_journal - carried_count[journal] for journal in journals}
    remaining_pool = [r for r in deduped if str(r.get("doi", "")).strip() not in prev_dois]
    new_rows = sample_articles_from_journal_quota(
        remaining_pool, quota=residual_quota, seed=seed + human_number,
        on_shortfall=on_shortfall,
        # Last-resort fallback for backfill mode: the FULL eligible pool
        # (including articles the previous tester already saw), consulted
        # only once `remaining_pool` itself can't cover the shortfall.
        backfill_pool=deduped,
    )

    # Re-interleave the carried + new articles together as one set, so
    # sibling summaries of one article are never adjacent regardless of
    # whether that article came from the carry or the fresh draw.
    combined_groups = _group_rows_by_article(carried_rows) + _group_rows_by_article(new_rows)
    selected = _interleave_article_groups(combined_groups, rng=random.Random(seed + human_number))
    return selected, sum(carried_count.values())


# ---------------------------------------------------------------------------
# Export entry point
# ---------------------------------------------------------------------------
# In plain English: everything above this line is a building block. The
# function below, ``export_human_review``, is the one function that ties
# them all together into the full "export" workflow described at the top of
# this file — this is what actually gets called (directly, or via the CLI
# at the bottom of the file) to produce ONE MORE humanN/ reviewer folder.

@dataclass(frozen=True)
class ExportResult:
    """A labeled form summarizing what one export run produced — handed back
    to whoever called ``export_human_review`` so they can report or inspect
    the results without re-reading files from disk.
    """

    human_number: int  # which tester this was (1, 2, 3, ...)
    human_dir: Path  # folder that was written, e.g. data/human_review/human1
    sample_size_requested: int
    items_exported: int
    overlap_units: int  # how many articles were carried from the previous tester
    skipped_rows: int
    unblinding_key_path: Path


def export_human_review(
    *,
    sample_size: int = MIN_ARTICLES,
    overlap_ratio: float = 0.6,
    seed: int = 42,
    evaluations_path: Path | None = None,
    output_dir: Path | None = None,
    summaries_path: Path | None = None,
    summaries_txt_dir: Path | None = None,
    dev_tests_summaries_txt_dir: Path | None = None,
    raw_dir: Path | None = None,
) -> ExportResult:
    """Sample evaluated (paper, summary) pairs and export ONE MORE blind reviewer packet.

    Each call writes the *next* ``humanN/`` folder under ``output_dir`` (never
    touching an earlier one), so real reviewers can be recruited and onboarded
    one at a time rather than all at once. Comparability between testers comes
    from ``overlap_ratio``: a fraction of the previous tester's articles (kept
    balanced per journal, see ``select_incremental_rows``) is carried into the
    new folder so Krippendorff's alpha has a shared subset to compute
    agreement over; the rest is a fresh draw.

    Draws an EXACT number of articles per journal via
    ``sample_articles_from_journal_quota`` / ``select_incremental_rows`` —
    ``sample_size`` must be a multiple of ``len(STUDY_JOURNALS)`` (5, 10, or
    25 for this study's 5 journals) or a
    :class:`JournalQuotaError`/``ValueError`` is raised rather than silently
    breaking the "equal per journal" guarantee. (The lower-level, free-form
    ``sample_rows_for_review`` sampler this used to delegate to is still
    available for direct/scripted use — it's just no longer what this
    export function calls.)

    In plain English: this is the main "do the export" function — call this
    (or run the command-line tool at the bottom of this file, which calls
    it) to generate the next reviewer folder from whatever is currently in
    data/evaluations.jsonl.
    """
    if sample_size < 1:
        raise ValueError(f"sample_size must be >= 1, got {sample_size}")

    # Fall back to this module's default data/ locations for any path the
    # caller didn't explicitly override.
    resolved_evaluations = evaluations_path if evaluations_path is not None else EVALUATIONS_PATH
    resolved_output = output_dir if output_dir is not None else HUMAN_REVIEW_DIR
    resolved_summaries = summaries_path if summaries_path is not None else SUMMARIES_PATH
    resolved_txt_dir = summaries_txt_dir if summaries_txt_dir is not None else SUMMARIES_TXT_DIR
    resolved_dev_txt_dir = (
        dev_tests_summaries_txt_dir if dev_tests_summaries_txt_dir is not None
        else DEV_TESTS_SUMMARIES_TXT_DIR
    )
    resolved_raw = raw_dir if raw_dir is not None else RAW_DIR

    rows = list(iter_evaluation_rows(resolved_evaluations))

    # Step 1: figure out which tester this run is, and pick their rows.
    human_number = next_human_number(resolved_output)
    previous_key = load_previous_unblinding_key(resolved_output, human_number)
    sampled_rows, overlap_units = select_incremental_rows(
        rows, human_number=human_number, sample_size=sample_size, seed=seed,
        overlap_ratio=overlap_ratio, previous_key=previous_key, on_shortfall="error",
    )

    # Step 2: fetch the full article/summary text for each sampled row and
    # bundle it into ReviewItem objects (see _build_instance_lookup /
    # _build_review_items above).
    lookup = _build_instance_lookup(resolved_summaries, resolved_txt_dir, resolved_dev_txt_dir)
    items, skipped_rows = _build_review_items(sampled_rows, lookup)

    if skipped_rows:
        print(f"[human_review] WARNING: {len(skipped_rows)} sampled row(s) could not be "
              "matched back to source text (data/summaries.jsonl or the summarize-all "
              ".txt folders may have moved or been regenerated since evaluation) and "
              "were excluded from the export.")

    # Step 3: write the one humanN/ folder for this run (guide + packet index
    # + .xlsx scoresheet + one item_id/ folder per item + matched PDFs).
    resolved_output.mkdir(parents=True, exist_ok=True)
    human_dir = resolved_output / f"human{human_number}"
    scoresheet_name = f"scoresheet_human{human_number}.xlsx"
    _copied, pdfs_missing = write_review_folder(
        items, human_dir,
        reviewer_id=human_number,
        scoresheet_filename=scoresheet_name,
        raw_dir=resolved_raw,
        lookup=lookup,
    )

    if pdfs_missing:
        print(f"[human_review] WARNING: no source PDF found in {resolved_raw} for "
              f"{len(set(pdfs_missing))} article(s): " + ", ".join(sorted(set(pdfs_missing)))
              + ". The packet/article text still covers them; only the supplementary "
              "PDF is absent.")

    # Step 4: write this tester's private answer-key file (a SIBLING of
    # humanN/, never inside it), and print a summary + a loud reminder not
    # to share it with reviewers.
    key = build_unblinding_key(items, seed=seed, sample_size=sample_size, reviewer_count=1)
    key["human_number"] = human_number
    key["overlap_units"] = overlap_units
    key["overlap_ratio"] = overlap_ratio
    key["sample_unit"] = "articles"
    key_path = unblinding_key_path_for(resolved_output, human_number)
    key_path.write_text(json.dumps(key, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"[human_review] exported human{human_number} ({len(items)} item(s), "
          f"{overlap_units} article(s) carried from human{human_number - 1 if human_number > 1 else '-'}) "
          f"to {human_dir}")
    print(f"[human_review] un-blinding key written to {key_path} — "
          "DO NOT share this file with reviewers.")

    return ExportResult(
        human_number=human_number,
        human_dir=human_dir,
        sample_size_requested=sample_size,
        items_exported=len(items),
        overlap_units=overlap_units,
        skipped_rows=len(skipped_rows),
        unblinding_key_path=key_path,
    )


# ===========================================================================
# INGEST — filled scoresheets -> normalized data/human_reviews.jsonl
# ===========================================================================
# In plain English: this second half of the file is the "read the reviewers'
# answers back in" half. A human reviewer has now opened their
# humanN/ folder, read the articles and summaries, and typed 1-5 scores
# into their scoresheet .xlsx file. The functions below find those filled
# scoresheets, read each row, look up that reviewer's own private unblinding
# key to find out which AI actually wrote each summary, and save everything
# into one tidy, combined file: data/human_reviews.jsonl. That file is what later code uses
# to check "did the human reviewers agree with the AI judges?"

def _reviewer_id_from_path(sheet_path: Path) -> str:
    """Best-effort reviewer id: the ``humanN`` folder, else the filename stem.

    Export writes ``humanN/scoresheet_humanN.xlsx`` (both the real study
    export and the pilot); a returned sheet normally keeps that layout, so
    the parent folder name is the reliable id — and it's also the exact key
    ``load_unblinding_keys`` uses to look up which key file un-blinds this
    reviewer's rows. A sheet handed back loosely (renamed, moved to the top
    level) falls back to its filename so a stray file is still attributable
    to *someone*, even though it then can't be un-blinded (see
    ``ingest_human_reviews``).

    In plain English: given the path to a filled-in scoresheet file, work
    out a short label for "which reviewer filled this in" — normally just
    the name of its containing folder (e.g. "human2").
    """
    parent = sheet_path.parent.name
    if parent.startswith("human"):
        return parent
    stem = sheet_path.stem  # e.g. scoresheet_human2
    marker = "scoresheet_"
    if stem.startswith(marker):
        return stem[len(marker):]
    return stem


def discover_scoresheets(review_dir: Path) -> list[Path]:
    """Find every filled scoresheet under a human-review export directory.

    The export writes ``.xlsx`` scoresheets under ``human*/scoresheet_human*.xlsx``
    (both the real study export and the pilot); ``.csv`` is still discovered
    so a reviewer who exported to CSV still ingests. As a fallback, a
    scoresheet returned loose at the top level (``scoresheet_*.{xlsx,csv}``)
    is picked up too. Sorted for deterministic ingest order.

    In plain English: scans the given folder for anything that looks like a
    filled-in scoresheet, checking a couple of possible filename patterns so
    ingest still works whether a reviewer kept the folder structure exactly
    as exported or just handed back a single file. ``review_dir.glob(pattern)``
    is a standard way to search a folder for files matching a wildcard
    pattern, similar to typing ``*.csv`` in a file browser's search box.
    """
    found: set[Path] = set()
    for pattern in (
        "human*/scoresheet_human*.xlsx",
        "human*/scoresheet_human*.csv",
        "scoresheet_*.xlsx",
        "scoresheet_*.csv",
    ):
        found.update(review_dir.glob(pattern))
    return sorted(found)


def _read_scoresheet_rows(path: Path) -> list[dict[str, str]]:
    """Read a filled scoresheet (``.xlsx`` or ``.csv``) as a list of row dicts.

    Both branches yield dicts keyed by ``SCORESHEET_FIELDS`` so
    ``_normalize_scoresheet_row`` handles either identically. The ``.xlsx``
    branch reads the header row and maps each subsequent row's cells to those
    headers (blank cells -> ""); ``openpyxl`` is imported lazily so ingest of a
    pure-CSV directory never needs it. The ``.csv`` branch uses the same
    ``utf-8-sig`` DictReader as before, so legacy sheets are unchanged.

    In plain English: reads one scoresheet file — whether it's an Excel
    .xlsx file or a plain .csv file — and turns it into the same shape of
    data either way: a list of dictionaries, one per row, each mapping a
    column name (like "faithfulness") to whatever the reviewer typed in
    that cell. That way every other function in this file can work with
    "a list of row dictionaries" without caring which file format it
    originally came from.
    """
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - exercised only without openpyxl
            raise RuntimeError(
                "Reading an .xlsx scoresheet needs the 'openpyxl' package "
                "(pip install openpyxl)."
            ) from exc
        # Open the workbook read-only (we're only reading it, not editing),
        # and pull out every row as plain values (not formula objects).
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            # The first row is the column headers (item_id, faithfulness, ...).
            header = [str(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            return []  # an empty sheet with no rows at all
        records: list[dict[str, str]] = []
        for raw in rows_iter:
            # Pad a short row with blanks in case Excel trimmed trailing
            # empty cells, then zip each cell up with its column header —
            # this is a dictionary comprehension: a compact way to build a
            # dictionary in one line by looping and computing each
            # key/value pair, instead of writing out a multi-line loop with
            # explicit .append()/assignment calls.
            cells = list(raw) + [None] * (len(header) - len(raw))
            records.append({
                header[i]: ("" if cells[i] is None else str(cells[i]))
                for i in range(len(header)) if header[i]
            })
        wb.close()
        return records

    # Plain CSV path: utf-8-sig strips a possible byte-order-mark that
    # Excel sometimes adds at the start of a CSV file it saved.
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_unblinding_keys(review_dir: Path) -> dict[str, dict[str, Any]]:
    """Load every ``unblinding_key_human{N}.json`` sibling, keyed by ``human{N}``.

    Each ``humanN/`` folder has its OWN private key (a sibling file, never
    inside the folder), and each key independently numbers its items starting
    at ``item_001``. Returning a ``{reviewer_id: key}`` mapping — instead of
    merging every key's ``items`` into one flat ``{item_id: identity}`` dict
    — is deliberate: ``human2``'s ``item_001`` and ``human1``'s ``item_001``
    describe two DIFFERENT (doi, summarizer) pairs, so a merged lookup would
    silently misattribute one reviewer's scores to another reviewer's
    identity the moment more than one humanN folder exists. ``ingest_human_reviews``
    looks up each scoresheet's rows against its OWN reviewer's key only.

    In plain English: opens every "answer key" file in the export folder —
    one per reviewer — and hands back a dictionary of dictionaries, so later
    code can say "give me human2's key" and get exactly that reviewer's
    identity mapping, never someone else's.
    """
    keys: dict[str, dict[str, Any]] = {}
    for key_path in sorted(review_dir.glob("unblinding_key_human*.json")):
        m = re.match(r"^unblinding_key_(human\d+)\.json$", key_path.name)
        if not m:
            continue
        try:
            keys[m.group(1)] = json.loads(key_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
    if not keys:
        raise FileNotFoundError(
            f"No unblinding_key_human*.json found in {review_dir}. Ingest needs the "
            "private un-blinding key(s) that 'export-human-review' wrote alongside "
            "each humanN/ folder; point --review-dir at that export directory."
        )
    return keys


def _parse_score(raw: str) -> tuple[float | None, bool]:
    """Parse one 1-5 criterion cell. Returns (value, was_invalid).

    Blank -> (None, False): the reviewer legitimately left it empty. A value
    outside 1-5 or non-numeric -> (None, True): flagged invalid so the caller
    can warn instead of silently trusting a typo like '44' or 'good'.

    In plain English: takes whatever raw text was typed into one score cell
    and tries to make sense of it. Returns two things back at once (a
    "tuple", a small fixed-size bundle of values): the cleaned-up number (or
    ``None`` if there wasn't a usable one), and a True/False flag saying
    whether what was typed looked like a mistake (so the caller can warn
    about it) rather than a deliberately-left-blank cell.
    """
    text = (raw or "").strip()
    if not text:
        return None, False
    try:
        value = float(text)
    except ValueError:
        return None, True
    if not (1.0 <= value <= 5.0):
        return None, True
    return value, False


def _parse_bool(raw: str) -> bool | None:
    """Parse a yes/no hallucination cell; blank or unrecognized -> None.

    In plain English: turns whatever text was typed in the
    hallucination_present cell (e.g. "yes", "Y", "no", "0") into a proper
    True/False, accepting several common spellings; anything left blank or
    that doesn't match a recognized word comes back as ``None`` (unknown).
    """
    text = (raw or "").strip().lower()
    if text in {"yes", "y", "true", "1"}:
        return True
    if text in {"no", "n", "false", "0"}:
        return False
    return None


def _normalize_scoresheet_row(
    raw_row: dict[str, str], reviewer_id: str, key_items: dict[str, Any],
) -> tuple[dict[str, Any] | None, list[str]]:
    """Turn one filled CSV row into a normalized human-review record.

    Returns (record, warnings). ``record`` is None when the row is blank (no
    scores, no hallucination flag, no free text) or when its ``item_id`` is not
    in the un-blinding key (can't attribute it to a summary). Un-blinded fields
    (doi/summariser/LLM scores) are copied from the key so the normalized row
    is self-contained for later correlation without re-reading it.

    In plain English: this is the heart of ingest — it takes ONE raw row
    from a reviewer's filled-in spreadsheet and turns it into one clean,
    validated, "un-blinded" record. Several checks happen in sequence, each
    one able to bail out early (returning ``None`` for the record) if the
    row isn't usable:
    1. Skip rows with no item_id at all.
    2. Parse and validate each of the five 1-5 criterion scores.
    3. Parse the yes/no hallucination flag and free-text notes/comment.
    4. If literally nothing was filled in, treat it as "not yet scored" and
       skip it quietly (not an error — just an empty row in the sheet).
    5. If something WAS filled in but the item_id doesn't match anything in
       the unblinding key, warn and skip it (can't attribute the score to
       any known summary).
    6. Otherwise, compute the human's overall scores and build the final
       record, filling in the real identity (doi, summarizer, ...) from the
       unblinding key.
    """
    warnings: list[str] = []
    # Step 1: no item_id -> nothing to key this row to; skip silently.
    item_id = (raw_row.get("item_id") or "").strip()
    if not item_id:
        return None, warnings

    # Step 2: parse each of the five 1-5 criterion cells. A malformed value
    # (not 1-5, not a number) is dropped but noted as a warning; a properly
    # blank cell is just dropped with no warning.
    criteria_scores: dict[str, float] = {}
    for criterion in CRITERIA:
        value, invalid = _parse_score(raw_row.get(criterion, ""))
        if invalid:
            warnings.append(
                f"{reviewer_id}/{item_id}: '{criterion}'={raw_row.get(criterion)!r} "
                "is not a 1-5 score; ignored."
            )
        elif value is not None:
            criteria_scores[criterion] = value

    # Step 3: parse the remaining free-form fields for this row.
    hallucination_present = _parse_bool(raw_row.get("hallucination_present", ""))
    notes = (raw_row.get("hallucination_notes") or "").strip()
    comment = (raw_row.get("comment") or "").strip()

    # Step 4: a row with nothing on it is a not-yet-scored item, not data — drop it
    # silently so a half-filled sheet ingests cleanly.
    if not criteria_scores and hallucination_present is None and not notes and not comment:
        return None, warnings

    # Step 5: the row has real data, but we can't identify what it's scoring.
    if item_id not in key_items:
        warnings.append(
            f"{reviewer_id}/{item_id}: no matching item in {reviewer_id}'s unblinding "
            "key; skipped."
        )
        return None, warnings

    # Step 6: we have valid scores and a known item_id — un-blind it (look
    # up its real identity in the key) and compute the human's composite
    # scores to build the final record.
    identity = key_items[item_id]
    human_score = (
        round(sum(criteria_scores.values()) / len(criteria_scores), 3)
        if criteria_scores else None
    )
    # Clinical-risk-weighted human composite, mirroring the jury's own
    # jury_score_weighted (calculate_jury_score with MEDHELM_CRITERION_WEIGHTS is
    # the single source of truth). Only when ALL five criteria are present:
    # calculate_jury_score clamps a missing criterion to 1, which would silently
    # deflate a partially-filled row, so a partial row gets None instead.
    human_score_weighted = (
        calculate_jury_score(criteria_scores, weights=MEDHELM_CRITERION_WEIGHTS)
        if len(criteria_scores) == len(CRITERIA) else None
    )
    record = {
        "item_id": item_id,
        "reviewer_id": reviewer_id,
        "doi": identity.get("doi", ""),
        "summarizer": identity.get("summarizer", ""),
        "judge": identity.get("judge", ""),
        "input_source": identity.get("input_source", "processed"),
        "rubric_version": identity.get("rubric_version", ""),
        "strata": identity.get("strata") or {},
        "criteria_scores": criteria_scores,
        "human_score_unweighted": human_score,
        "human_score_weighted": human_score_weighted,
        "hallucination_present": hallucination_present,
        "hallucination_notes": notes,
        "comment": comment,
        # LLM jury's own scores for this item, carried from the key so the
        # correlation step never needs to re-open evaluations.jsonl.
        "llm_jury_score": identity.get("llm_jury_score"),
        "llm_jury_score_weighted": identity.get("llm_jury_score_weighted"),
        "llm_jury_score_unweighted": identity.get("llm_jury_score_unweighted"),
        "llm_criteria_scores": identity.get("llm_criteria_scores") or {},
    }
    return record, warnings


@dataclass(frozen=True)
class IngestResult:
    """A labeled form summarizing what one ingest run did: where the output
    went, who the reviewers were, how many rows were written, and any
    warnings raised along the way — returned to the caller instead of
    forcing them to re-read the output file to find out what happened.
    """

    output_path: Path
    reviewers: list[str]
    rows_written: int
    scoresheets_read: int
    warnings: list[str]


def ingest_human_reviews(
    *,
    review_dir: Path | None = None,
    output_path: Path | None = None,
) -> IngestResult:
    """Read every filled scoresheet under ``review_dir`` into one JSONL file.

    Idempotent: the output is rewritten (not appended) each run, so ingesting
    the same sheets twice yields the same file. Rows are sorted by
    (item_id, reviewer_id) for a stable, diff-friendly artifact.

    In plain English: this is the main "do the ingest" function — the
    counterpart to ``export_human_review`` above. It finds every filled
    scoresheet under the export folder, cleans up and un-blinds every row
    (via ``_normalize_scoresheet_row``), and saves the results as one JSON
    Lines file (a text file with one JSON record per line — "JSONL" for
    short). "Idempotent" means running it twice with the same input files
    produces the exact same output — it overwrites the file fresh each time
    rather than appending duplicate copies, so it's always safe to re-run
    after fixing a typo in a scoresheet.
    """
    resolved_dir = review_dir if review_dir is not None else HUMAN_REVIEW_DIR
    resolved_output = output_path if output_path is not None else HUMAN_REVIEWS_PATH

    # One key PER reviewer (see load_unblinding_keys) — deliberately never
    # merged into a single flat {item_id: identity} dict, since human2's
    # item_001 and human1's item_001 identify different summaries.
    keys_by_reviewer = load_unblinding_keys(resolved_dir)

    scoresheets = discover_scoresheets(resolved_dir)
    records: list[dict[str, Any]] = []
    reviewers: set[str] = set()
    warnings: list[str] = []

    # For every scoresheet found, read its rows and normalize each one
    # against THAT reviewer's own key only; keep only the rows that came back
    # as real records (None means "blank or unusable, skip it" — see
    # _normalize_scoresheet_row).
    for sheet_path in scoresheets:
        reviewer_id = _reviewer_id_from_path(sheet_path)
        reviewers.add(reviewer_id)
        reviewer_key = keys_by_reviewer.get(reviewer_id)
        if reviewer_key is None:
            warnings.append(
                f"{reviewer_id}: no unblinding_key_{reviewer_id}.json found; every row "
                f"in {sheet_path.name} was skipped (can't un-blind without it)."
            )
            continue
        key_items = reviewer_key.get("items") or {}
        for raw_row in _read_scoresheet_rows(sheet_path):
            record, row_warnings = _normalize_scoresheet_row(raw_row, reviewer_id, key_items)
            warnings.extend(row_warnings)
            if record is not None:
                records.append(record)

    # Sort for a stable, predictable file — re-running ingest on unchanged
    # inputs always produces byte-identical output.
    records.sort(key=lambda r: (r["item_id"], r["reviewer_id"]))

    resolved_output.parent.mkdir(parents=True, exist_ok=True)
    with open(resolved_output, "w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    for warning in warnings:
        print(f"[human_review] WARNING: {warning}")
    print(f"[human_review] ingested {len(records)} scored row(s) from "
          f"{len(scoresheets)} scoresheet(s) ({len(reviewers)} reviewer(s)) -> {resolved_output}")

    return IngestResult(
        output_path=resolved_output,
        reviewers=sorted(reviewers),
        rows_written=len(records),
        scoresheets_read=len(scoresheets),
        warnings=warnings,
    )


def iter_human_review_rows(path: Path | None = None) -> Iterable[dict[str, Any]]:
    """Yield normalized human-review rows from data/human_reviews.jsonl.

    In plain English: opens the ingested human_reviews.jsonl file and, one
    line at a time, hands back each row as a Python dictionary. "Yield"
    (rather than "return") makes this a generator — it produces rows one at
    a time as they're asked for, instead of loading the whole file into
    memory as one big list up front. If the file doesn't exist yet (no
    reviews ingested), it simply produces nothing rather than erroring.
    """
    resolved = path if path is not None else HUMAN_REVIEWS_PATH
    if not resolved.exists():
        return
    with open(resolved, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(row, dict):
                yield row


# ===========================================================================
# ANALYSIS — inter-reviewer agreement + human-vs-jury correlation
# ===========================================================================
# In plain English: now that human_reviews.jsonl exists (built by the ingest
# functions above), this section answers the two questions the whole human
# validation study exists to answer:
#   1. Do the human reviewers agree with EACH OTHER? (only meaningful with
#      2+ reviewers scoring the same items)
#   2. Do the human reviewers agree with the AI JURY's scores? (meaningful
#      even with just one reviewer)
# Both questions reuse statistics helpers from reliability.py that were
# originally written to compare AI judges against each other — a human
# reviewer is just treated as one more "rater" in that same math.

def _llm_criterion_scores(llm_criteria: Any) -> dict[str, float]:
    """Flatten the LLM's nested {crit: {score: n}} block to {crit: n}.

    In plain English: the AI judges store each criterion's score alongside
    its written reasoning, like ``{"faithfulness": {"score": 4, "reasoning":
    "..."}}``. This strips that down to just the numbers,
    ``{"faithfulness": 4}``, which is all the correlation math below needs.
    """
    flat: dict[str, float] = {}
    if isinstance(llm_criteria, dict):
        for criterion, detail in llm_criteria.items():
            score = detail.get("score") if isinstance(detail, dict) else detail
            if isinstance(score, (int, float)):
                flat[str(criterion)] = float(score)
    return flat


def _relabel_agreement(rel: dict[str, Any], n_reviewers: int) -> dict[str, Any]:
    """Recast reliability.compute_reliability's judge-flavored dict for humans.

    The stats engine is identical (each reviewer is a "rater"), only the labels
    change: judges -> raters, jury_score -> overall. When it is unavailable
    because there is only one reviewer, the judge-specific hint (about
    JUDGE_MODELS) is replaced with a human-appropriate reason.

    In plain English: ``reliability.compute_reliability`` was written for
    comparing AI judges and returns a dictionary that talks about "judges"
    and "jury_score". This function takes that same dictionary and just
    renames a few keys/wording so the result reads naturally when it's
    actually about human reviewers instead — no numbers are recalculated,
    only relabeled.
    """
    out: dict[str, Any] = {
        "available": bool(rel.get("available")),
        "raters": rel.get("judges", []),
        "n_raters": rel.get("n_judges", 0),
        "n_comparable_items": rel.get("n_comparable_items", 0),
        "overall": rel.get("jury_score"),
        "per_criterion": rel.get("per_criterion", {}),
        "pairwise_agreement": rel.get("pairwise_agreement", {}),
    }
    if rel.get("interpretation"):
        out["interpretation"] = rel["interpretation"]
    if not out["available"]:
        if n_reviewers < 2:
            out["reason"] = (
                f"Inter-reviewer agreement needs at least two reviewers; this export "
                f"was scored by {n_reviewers} reviewer(s). Human-vs-jury correlation is "
                "still reported below."
            )
        else:
            out["reason"] = rel.get(
                "reason", "Two or more reviewers scored, but no shared item to compare."
            )
    return out


def _inter_reviewer_agreement(rows: list[dict[str, Any]], n_reviewers: int) -> dict[str, Any]:
    """Inter-reviewer Krippendorff's alpha, reusing the judge reliability engine.

    Each normalized human row is reshaped into the (doi, summarizer,
    input_source, judge=reviewer, jury_score, criteria_scores) form
    reliability.compute_reliability already understands, so a reviewer is
    treated exactly as a judge — the same alpha, per-criterion spread, and
    pairwise agreement, with no duplicated statistics code.

    In plain English: "Krippendorff's alpha" is a standard statistic that
    answers "how much do multiple raters agree with each other", on a scale
    that accounts for chance agreement (not just "they gave the same
    number"). This function reshapes each human review row into the exact
    same shape reliability.py already knows how to feed that statistic
    (originally built for AI judges), calling each reviewer a "judge" just
    for this calculation, then calls the shared engine and relabels the
    result back to human-friendly wording via ``_relabel_agreement``.
    """
    rater_rows = [
        {
            "doi": row.get("doi", ""),
            "summarizer": row.get("summarizer", ""),
            "input_source": row.get("input_source", "processed"),
            "judge": row.get("reviewer_id", ""),
            "jury_score": row.get("human_score_unweighted"),
            "criteria_scores": {c: {"score": v} for c, v in (row.get("criteria_scores") or {}).items()},
        }
        for row in rows
        if row.get("human_score_unweighted") is not None
    ]
    return _relabel_agreement(reliability.compute_reliability(rater_rows), n_reviewers)


def _correlate(human: list[float], jury: list[float]) -> dict[str, Any]:
    """Pearson + Spearman + Bland-Altman + Cohen's Kappa for one aligned
    human/jury series.

    In plain English: given two matching lists of numbers — one reviewer's
    scores and the AI jury's scores for the same items, in the same order —
    this runs a battery of standard statistics that answer "how closely do
    these two sets of numbers track each other?" from a few different
    angles (straight-line correlation, rank-order correlation, average bias,
    and exact-category agreement), and packages all the results into one
    dictionary.

    Pearson/Spearman/Bland-Altman measure rank/linear agreement on the raw
    continuous scores. Cohen's Kappa (+ percent agreement) is a different,
    categorical question — "how often do the two raters land on the exact
    same 1-5 category, correcting for the agreement expected by chance" — so
    both are reported rather than one replacing the other. Kappa needs
    discrete categories; stats_engine.categorical_agreement rounds the
    continuous composites to the nearest 1-5 integer first.
    """
    n = len(human)
    pearson = reliability.pearson(human, jury)
    spearman = reliability.spearman(human, jury)
    # Spearman is the headline (rank agreement); fall back to Pearson only for
    # the interpretation line when ranks are undefined (e.g. all-tied scores).
    headline = spearman if spearman is not None else pearson
    return {
        "n": n,
        "pearson": pearson,
        # Two-sided p-values (scipy) so a correlation can be reported as
        # "r=0.72, p=0.003" — the significance a validation claim needs, not just
        # the coefficient. None when undefined (e.g. p is not defined at n=2).
        "pearson_p": reliability.pearson_p(human, jury),
        "spearman": spearman,
        "spearman_p": reliability.spearman_p(human, jury),
        "bland_altman": reliability.bland_altman(human, jury),
        **categorical_agreement(human, jury),
        # n is passed so the interpretation withholds a verdict on an
        # underpowered sample (see reliability.MIN_CORRELATION_N).
        "interpretation": reliability.interpret_correlation(headline, n),
    }


def _human_vs_jury(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Correlate mean human scores against the LLM jury, overall + per-criterion.

    Human scores are averaged across whoever reviewed an item, then paired with
    that item's stored LLM jury score. Two overall correlations are produced so
    BOTH jury modes are validated against humans:

    - ``overall`` — unweighted composites on both sides (plain means of the five
      criteria). This is the MedHELM-comparable primary: MedHELM's own
      jury_score is an unweighted pooled mean (llm_jury_metrics.py).
    - ``overall_weighted`` — this project's clinical-risk-weighted composites
      (faithfulness/safety count more), so the safety-oriented score the study
      actually reports is validated too, not left unchecked.

    Per-criterion pairs each human criterion mean with the LLM's score for the
    same criterion (weighting is not meaningful for a single criterion).

    In plain English: this is the "does the AI jury track expert judgment?"
    calculation. For every scored item, it lines up "what did the human(s)
    say" against "what did the AI jury say" — for the overall score (both
    the plain-average and the clinical-risk-weighted version) and for each
    of the five criteria separately — then hands each pair of lists off to
    ``_correlate`` to get the actual statistics.
    """
    # Group by item so multi-reviewer scores collapse to one human value per item.
    overall_human: dict[str, list[float]] = defaultdict(list)
    overall_jury: dict[str, float] = {}
    overall_human_w: dict[str, list[float]] = defaultdict(list)
    overall_jury_w: dict[str, float] = {}
    crit_human: dict[str, dict[str, list[float]]] = {c: defaultdict(list) for c in CRITERIA}
    crit_jury: dict[str, dict[str, float]] = {c: {} for c in CRITERIA}

    # Walk every scored row once, sorting each value into the right bucket
    # above by item_id: human scores accumulate in a list per item (there
    # may be several reviewers), while the jury only ever has one score per
    # item so it's stored directly.
    for row in rows:
        item_id = row.get("item_id", "")
        human_overall = row.get("human_score_unweighted")
        if human_overall is not None:
            overall_human[item_id].append(float(human_overall))
        jury_overall = row.get("llm_jury_score_unweighted")
        if jury_overall is None:
            jury_overall = row.get("llm_jury_score")
        if isinstance(jury_overall, (int, float)):
            overall_jury[item_id] = float(jury_overall)

        human_overall_w = row.get("human_score_weighted")
        if human_overall_w is not None:
            overall_human_w[item_id].append(float(human_overall_w))
        jury_overall_w = row.get("llm_jury_score_weighted")
        if isinstance(jury_overall_w, (int, float)):
            overall_jury_w[item_id] = float(jury_overall_w)

        llm_crit = _llm_criterion_scores(row.get("llm_criteria_scores"))
        for criterion, value in (row.get("criteria_scores") or {}).items():
            if criterion in crit_human:
                crit_human[criterion][item_id].append(float(value))
        for criterion, score in llm_crit.items():
            if criterion in crit_jury:
                crit_jury[criterion][item_id] = score

    # A small helper function defined right here (a "nested function"),
    # since it's only ever used inside _human_vs_jury: turns the per-item
    # buckets built above into two same-length, same-order lists — one of
    # human scores, one of jury scores — keeping only items both sides
    # actually have a score for (an item with no human score, or no jury
    # score, can't be compared and is left out).
    def paired(human_by_item: dict[str, list[float]], jury_by_item: dict[str, float]):
        human_series: list[float] = []
        jury_series: list[float] = []
        for item_id, human_values in human_by_item.items():
            if item_id in jury_by_item and human_values:
                human_series.append(sum(human_values) / len(human_values))
                jury_series.append(jury_by_item[item_id])
        return human_series, jury_series

    overall_h, overall_j = paired(overall_human, overall_jury)
    overall_hw, overall_jw = paired(overall_human_w, overall_jury_w)
    per_criterion: dict[str, dict[str, Any]] = {}
    for criterion in CRITERIA:
        h, j = paired(crit_human[criterion], crit_jury[criterion])
        per_criterion[criterion] = _correlate(h, j)

    return {
        "overall": _correlate(overall_h, overall_j),
        "overall_weighted": _correlate(overall_hw, overall_jw),
        "per_criterion": per_criterion,
    }


def human_vs_jury_by_provider(rows: Iterable[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Human-vs-jury correlation, one result per summarizer provider.

    In plain English: runs the same "does the human agree with the jury?"
    calculation as above, but separately for each AI provider (OpenAI,
    Anthropic, Gemini, ...) instead of pooling every provider's rows
    together — so you can ask "is the jury just as trustworthy when judging
    OpenAI's summaries as when judging Anthropic's?"

    Same statistics as :func:`analyze_human_reviews`'s pooled view
    (:func:`_human_vs_jury`), but each provider's reviewed items are
    correlated on their own — answers "does the jury track expert judgment
    specifically for provider X's summaries?" rather than the pooled figure
    across all providers. Used by ``report_figures.py``'s leaderboard, which
    breaks every other column out per provider too.

    Providers with no scored rows are simply absent from the result. With a
    typical human-review sample size split three ways, each provider's ``n``
    is usually below ``reliability.MIN_CORRELATION_N`` — the coefficient is
    still returned (callers see the interpretation's underpowered caveat via
    the same ``interpret_correlation`` gate ``_human_vs_jury`` already applies).
    """
    scored = [r for r in rows if isinstance(r, dict) and (
        r.get("human_score_unweighted") is not None or (r.get("criteria_scores") or {})
    )]
    providers = sorted({str(r.get("summarizer", "")).strip() for r in scored if r.get("summarizer")})
    return {
        provider: _human_vs_jury([r for r in scored if str(r.get("summarizer", "")).strip() == provider])
        for provider in providers
    }


# How human-vs-jury correlation is broken out. See docs/phase5/human_validation.md.
# In plain English: the three allowed settings for "mode" wherever this file
# mentions human-validation mode — see analyze_human_reviews's docstring
# below for what each one means.
HUMAN_VALIDATION_MODES = ("per_reviewer", "pooled", "both")
DEFAULT_HUMAN_VALIDATION_MODE = "per_reviewer"


def resolve_human_validation_mode(mode: str | None) -> str:
    """Normalize a human-validation mode, defaulting safely.

    Precedence is the caller's job (CLI > env > default); this only validates.
    An unknown value falls back to the default with a warning rather than
    raising, so a typo in .env never crashes an offline report.

    In plain English: takes whatever "mode" text was configured (which might
    be missing, or contain a typo) and returns one of the three valid modes
    — printing a warning and using the safe default instead of crashing if
    what was given doesn't match anything recognized.
    """
    resolved = (mode or DEFAULT_HUMAN_VALIDATION_MODE).strip().lower()
    if resolved not in HUMAN_VALIDATION_MODES:
        print(f"[human_review] WARNING: invalid human-validation mode {mode!r}; "
              f"using {DEFAULT_HUMAN_VALIDATION_MODE!r}. Valid: {HUMAN_VALIDATION_MODES}")
        return DEFAULT_HUMAN_VALIDATION_MODE
    return resolved


def analyze_human_reviews(
    rows: Iterable[dict[str, Any]], *, mode: str | None = None,
) -> dict[str, Any]:
    """Full human-validation view: inter-reviewer agreement + human-vs-jury.

    ``mode`` controls how human-vs-jury correlation is reported:

    - ``per_reviewer`` (default) — a separate correlation per reviewer, so an
      expert (e.g. a veterinarian) is validated against the jury on their OWN
      scores, never diluted by averaging with a non-expert reviewer. This is
      what a claim like "the jury tracks veterinarian judgment" actually needs.
    - ``pooled`` — one correlation over the mean human score per item (all
      reviewers averaged together). Simpler, but a mixed-expertise panel makes
      "the human" an average of expert and non-expert.
    - ``both`` — report both views.

    Always returns the same top-level shape; ``available`` is False (with a
    ``reason``) when there are no scored human rows yet, so eval_report can
    render the section without special-casing missing keys — mirroring
    reliability.compute_reliability's contract.

    In plain English: this is the main "analyze the human review results"
    function — the counterpart to ``export_human_review`` (build the study)
    and ``ingest_human_reviews`` (collect the answers). Feed it the rows
    from ``iter_human_review_rows()`` and it produces one dictionary
    containing everything needed for a report: how many reviewers/items
    there were, whether reviewers agreed with each other, and whether they
    agreed with the AI jury. It always returns the same set of keys (with
    ``available: False`` and a plain-English ``reason`` when there's nothing
    to analyze yet) so code that displays this doesn't need special-case
    handling for "no data yet".
    """
    resolved_mode = resolve_human_validation_mode(mode)
    materialized = [r for r in rows if isinstance(r, dict)]
    scored = [r for r in materialized if r.get("human_score_unweighted") is not None
              or (r.get("criteria_scores") or {})]
    reviewers = sorted({str(r.get("reviewer_id", "")) for r in materialized if r.get("reviewer_id")})
    items = sorted({str(r.get("item_id", "")) for r in scored if r.get("item_id")})

    # Nothing has been scored yet — return the "not available" shape rather
    # than trying (and failing) to compute statistics on an empty dataset.
    if not scored:
        return {
            "available": False,
            "reason": (
                "No human_reviews.jsonl rows yet. Run 'export-human-review', have "
                "reviewers fill the scoresheets, then 'ingest-human-review'."
            ),
            "mode": resolved_mode,
            "n_reviewers": len(reviewers),
            "reviewers": reviewers,
            "n_items": 0,
            "inter_reviewer_agreement": {"available": False, "reason": "No scored rows."},
            "human_vs_jury": None,
            "by_reviewer": {},
        }

    result: dict[str, Any] = {
        "available": True,
        "mode": resolved_mode,
        "n_reviewers": len(reviewers),
        "reviewers": reviewers,
        "n_items": len(items),
        "inter_reviewer_agreement": _inter_reviewer_agreement(scored, len(reviewers)),
        "human_vs_jury": None,
        "by_reviewer": {},
    }

    if resolved_mode in ("pooled", "both"):
        result["human_vs_jury"] = _human_vs_jury(scored)
    if resolved_mode in ("per_reviewer", "both"):
        # _human_vs_jury on one reviewer's rows: with a single rater per item
        # the "average across reviewers" collapses to that reviewer's score, so
        # the same tested function computes each reviewer's correlation cleanly.
        result["by_reviewer"] = {
            reviewer_id: _human_vs_jury([r for r in scored if r.get("reviewer_id") == reviewer_id])
            for reviewer_id in reviewers
        }

    # A single top-line interpretation only makes sense for the pooled view;
    # in per_reviewer mode each reviewer carries its own (there is no single
    # "human"), so the renderer reads per-block interpretations instead.
    pooled = result.get("human_vs_jury")
    result["interpretation"] = (
        (pooled.get("overall") or {}).get("interpretation") if pooled else None
    )
    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
# In plain English: "CLI" stands for command-line interface — this section is
# what lets a person type a command like ``python human_review.py
# --overlap-ratio 0.6`` in a terminal and have it run the export (or, in
# ``ingest_main``, the ingest) with the settings they specified. It reads
# settings from three possible places, in priority order: an explicit
# command-line flag, then a setting in the .env configuration file, then a
# hard-coded default.

def _env_int(name: str, default: int) -> int:
    """Read an integer setting from an environment variable (e.g. from .env),
    falling back to ``default`` if it's missing or not a valid whole number
    (printing a warning in the latter case rather than crashing).
    """
    raw = os.getenv(name)
    if raw is None or not raw.strip():
        return default
    try:
        return int(raw)
    except ValueError:
        print(f"[human_review] WARNING: invalid {name}={raw!r}; using {default}.")
        return default


def _env_float(name: str, default: float) -> float:
    """Same as :func:`_env_int` but for decimal-number settings (e.g. a ratio)."""
    raw = os.getenv(name)
    if raw is None or not raw.strip():
        return default
    try:
        return float(raw)
    except ValueError:
        print(f"[human_review] WARNING: invalid {name}={raw!r}; using {default}.")
        return default


def _clamp_ratio(value: float) -> float:
    """Keep the overlap ratio in [0, 1] with a warning if it was out of range.

    In plain English: the overlap ratio only makes sense between 0 (nothing
    carried over) and 1 (everything carried over). If someone passes
    something outside that range, this quietly pulls it back into range and
    prints a warning, instead of letting a nonsensical value through.
    """
    if value < 0.0 or value > 1.0:
        clamped = max(0.0, min(1.0, value))
        print(f"[human_review] WARNING: overlap ratio {value} out of [0, 1]; "
              f"using {clamped}.")
        return clamped
    return value


SAMPLE_UNITS = ("items", "articles")
# Default to "articles": the reviewer folder is built around N articles (one per
# journal), each paired with every provider's summary. The low-level sampler
# (sample_rows_for_review) keeps its own "items" default for callers that want a
# single (article, provider) pair per count; only the export CLIs default here.
DEFAULT_SAMPLE_UNIT = "articles"


def resolve_sample_unit(unit: str | None) -> str:
    """Normalize ``sample_unit``, defaulting safely (mirrors resolve_human_validation_mode).

    An unknown value falls back to the default with a warning rather than
    raising, so a typo in .env never crashes an offline export.

    In plain English: same idea as ``resolve_human_validation_mode`` above,
    but for the "items" vs "articles" sampling-unit setting instead.
    """
    resolved = (unit or DEFAULT_SAMPLE_UNIT).strip().lower()
    if resolved not in SAMPLE_UNITS:
        print(f"[human_review] WARNING: invalid sample unit {unit!r}; "
              f"using {DEFAULT_SAMPLE_UNIT!r}. Valid: {SAMPLE_UNITS}")
        return DEFAULT_SAMPLE_UNIT
    return resolved


def prompt_article_count(default: int = MIN_ARTICLES, *, input_fn=input, stream=None) -> int:
    """Interactively ask how many articles the reviewer will evaluate (min MIN_ARTICLES).

    Used by both exports' CLIs only when no ``--sample-size`` was passed, so the
    unit-tested export functions themselves never block on input. Re-asks on a
    non-numeric answer or a number below ``MIN_ARTICLES`` (one article per
    journal). When stdin is not a TTY — piped input, CI, the test suite — there
    is no one to prompt, so it falls back to ``max(default, MIN_ARTICLES)``
    instead of hanging.

    In plain English: if you run the export command without saying how many
    articles you want, and you're typing directly into a terminal (a "TTY"),
    this politely asks you and keeps re-asking until you give a valid whole
    number that's at least MIN_ARTICLES. If there's no human at the keyboard
    to answer (e.g. input piped from a script, or an automated test), it
    just picks the minimum sensible number instead of freezing forever
    waiting for an answer that will never come.
    """
    s = stream if stream is not None else sys.stdin
    floor = max(default, MIN_ARTICLES)
    if not getattr(s, "isatty", lambda: False)():
        return floor
    # Keep asking until we get a usable answer (a whole number >= MIN_ARTICLES)
    # or the user gives up (an empty answer or Ctrl-D/EOF), each of which
    # returns a safe fallback value instead of looping forever.
    while True:
        try:
            raw = input_fn(
                f"How many articles will you evaluate? (minimum {MIN_ARTICLES}, "
                "one per journal): "
            ).strip()
        except EOFError:
            return floor
        if not raw:
            return floor
        try:
            value = int(raw)
        except ValueError:
            print(f"[human_review] Please enter a whole number >= {MIN_ARTICLES}.")
            continue
        if value < MIN_ARTICLES:
            print(f"[human_review] Minimum is {MIN_ARTICLES} (one article per journal). "
                  "Please try again.")
            continue
        return value


def prompt_article_count_choice(
    default: int = ARTICLE_COUNT_CHOICES[0], *, input_fn=input, stream=None,
) -> int:
    """Interactively ask the reviewer to pick 5, 10, or 25 articles — nothing else.

    Used by both exports' CLIs (the ``sample_unit="articles"`` default flow)
    when no ``--sample-size`` was passed. Unlike ``prompt_article_count``
    (any whole number >= MIN_ARTICLES), this only accepts one of
    :data:`ARTICLE_COUNT_CHOICES` — typed either as the article count itself
    (``5``/``10``/``25``) or as the menu number (``1``/``2``/``3``) — because
    each choice must divide evenly across :data:`STUDY_JOURNALS` for the
    equal-per-journal guarantee to hold. Re-asks on anything else. When
    stdin is not a TTY — piped input, CI, the test suite — there is no one
    to prompt, so it falls back to ``default`` instead of hanging.

    In plain English: shows a short menu (5, 10, or 25 articles, and how
    many summaries that means reading) and keeps asking until you pick one
    of the three. If there's no human at the keyboard to answer, it just
    picks the smallest option instead of freezing forever.
    """
    s = stream if stream is not None else sys.stdin
    if default not in ARTICLE_COUNT_CHOICES:
        default = ARTICLE_COUNT_CHOICES[0]
    if not getattr(s, "isatty", lambda: False)():
        return default

    n_journals = len(STUDY_JOURNALS)
    menu_lines = [
        "This export samples articles evenly across the study's "
        f"{n_journals} journals ({', '.join(STUDY_JOURNALS)}).",
        "How many articles will you evaluate?",
    ]
    for i, choice in enumerate(ARTICLE_COUNT_CHOICES, start=1):
        menu_lines.append(
            f"  {i}) {choice} articles -> {choice * 3} summaries to read "
            f"({choice // n_journals} per journal)"
        )
    print("\n".join(menu_lines))

    choice_by_menu_number = {str(i): c for i, c in enumerate(ARTICLE_COUNT_CHOICES, start=1)}
    choice_strings = {str(c) for c in ARTICLE_COUNT_CHOICES}
    while True:
        try:
            raw = input_fn(
                "Enter 1, 2, 3, or the article count directly "
                f"({'/'.join(str(c) for c in ARTICLE_COUNT_CHOICES)}): "
            ).strip()
        except EOFError:
            return default
        if not raw:
            return default
        if raw in choice_by_menu_number:
            return choice_by_menu_number[raw]
        if raw in choice_strings:
            return int(raw)
        print(f"[human_review] Please enter 1, 2, 3, or one of "
              f"{ARTICLE_COUNT_CHOICES}.")


def main(argv: list[str] | None = None) -> int:
    """Command-line entry point for the EXPORT half of this file.

    In plain English: this is what runs when someone types
    ``python human_review.py`` (with optional flags like ``--overlap-ratio``)
    in a terminal. Each call exports ONE MORE ``humanN/`` folder. It reads
    the command-line flags (falling back to .env settings, then hard-coded
    defaults), calls ``export_human_review`` with them, and returns an exit
    code — 0 for success, 1 for failure — which is the standard way a
    command-line program reports "did it work?" back to the terminal/scripts
    that ran it.
    """
    # argparse.ArgumentParser declares which command-line flags this program
    # accepts (e.g. --overlap-ratio, --sample-size); each add_argument() call
    # below describes one flag, its type, and its help text. None of these
    # flags are required — every one has a fallback below (.env, then a
    # built-in default).
    parser = argparse.ArgumentParser(
        description="Phase 5 — export ONE MORE blind human-validation reviewer "
                     "folder (humanN/) from data/evaluations.jsonl.",
    )
    parser.add_argument("--sample-size", type=int, default=None,
                        help="Articles to sample (must be a multiple of the number "
                             "of study journals, e.g. 5/10/25). Default: "
                             "HUMAN_REVIEW_SAMPLE_SIZE from .env, or an interactive "
                             "5/10/25 prompt at a real terminal.")
    parser.add_argument("--overlap-ratio", type=float, default=None,
                        help="Fraction of the previous tester's articles to carry "
                             "over, balanced per journal (default: "
                             "HUMAN_REVIEW_OVERLAP_RATIO, or 0.6). 1.0 = identical "
                             "items every run, 0.0 = a fresh draw each run.")
    parser.add_argument("--seed", type=int, default=None,
                        help="Sampling seed (default: HUMAN_REVIEW_SEED from .env, or 42).")
    parser.add_argument("--evaluations", type=Path, default=None,
                        help="Path to evaluations.jsonl (default: data/evaluations.jsonl).")
    parser.add_argument("--output-dir", type=Path, default=None,
                        help="Where to write reviewer folders (default: data/human_review/).")
    args = parser.parse_args(argv)

    # Resolve each setting: explicit flag wins if given, otherwise fall back
    # to the matching .env variable, otherwise use a built-in default.
    overlap_ratio = _clamp_ratio(
        args.overlap_ratio if args.overlap_ratio is not None
        else _env_float("HUMAN_REVIEW_OVERLAP_RATIO", 0.6)
    )
    seed = args.seed if args.seed is not None else _env_int("HUMAN_REVIEW_SEED", 42)
    # Sample size precedence: --sample-size > HUMAN_REVIEW_SAMPLE_SIZE > interactive
    # 5/10/25 prompt. The prompt only fires when nothing was configured AND
    # stdin is a TTY; otherwise it falls back to the smallest choice (see
    # prompt_article_count_choice), so scripted/CI runs never block.
    env_sample_size = _env_int("HUMAN_REVIEW_SAMPLE_SIZE", 0)
    if args.sample_size is not None:
        sample_size = args.sample_size
    elif env_sample_size > 0:
        sample_size = env_sample_size
    else:
        sample_size = prompt_article_count_choice(ARTICLE_COUNT_CHOICES[0])

    try:
        result = export_human_review(
            sample_size=sample_size,
            overlap_ratio=overlap_ratio,
            seed=seed,
            evaluations_path=args.evaluations,
            output_dir=args.output_dir,
        )
    except JournalQuotaError as exc:
        # The evaluated pool can't supply this run's exact per-journal quota
        # — report it plainly instead of a raw traceback.
        print(f"[human_review] {exc}")
        return 1
    # Exit code 0 = success, 1 = failure — the standard convention command-line
    # tools use so scripts/CI can check "did this succeed?" automatically.
    if result.items_exported == 0:
        print("[human_review] Nothing exported — run 'evaluate' first so "
              "data/evaluations.jsonl has rows to sample from.")
        return 1
    return 0


def ingest_main(argv: list[str] | None = None) -> int:
    """CLI for the ingest step (kept separate from the export ``main()``).

    Reads filled scoresheets from a human-review export directory into
    data/human_reviews.jsonl. eval-report then surfaces the human-validation
    analysis automatically when that file exists.

    In plain English: the command-line entry point for the INGEST half —
    what runs when someone types a command to ingest filled scoresheets
    (see docs/phase5/human_validation.md for the exact command). Reads its
    flags, calls ``ingest_human_reviews``, and returns a 0/1 exit code the
    same way ``main()`` does for export.
    """
    parser = argparse.ArgumentParser(
        description="Phase 5 — ingest filled human-validation scoresheets into "
                     "data/human_reviews.jsonl.",
    )
    parser.add_argument("--review-dir", type=Path, default=None,
                        help="Export directory holding humanN/ folders and their "
                             "sibling unblinding_key_human*.json files "
                             "(default: data/human_review/).")
    parser.add_argument("--output", type=Path, default=None,
                        help="Where to write the normalized JSONL "
                             "(default: data/human_reviews.jsonl).")
    args = parser.parse_args(argv)

    try:
        result = ingest_human_reviews(review_dir=args.review_dir, output_path=args.output)
    except FileNotFoundError as exc:
        # No unblinding_key_human*.json in the given folder — report it
        # plainly instead of showing a raw Python traceback.
        print(f"[human_review] {exc}")
        return 1

    if result.rows_written == 0:
        print("[human_review] No scored rows found — have reviewers fill in the "
              "scoresheet CSVs (1-5 per criterion) before ingesting.")
        return 1
    return 0


# In plain English: this only runs when the file is executed directly as a
# script (``python human_review.py ...``), not when it's imported by another
# file (e.g. by the test suite). It calls the export CLI and passes its
# 0/1 exit code back to the operating system via sys.exit — that's how a
# terminal or a script knows whether the command succeeded.
if __name__ == "__main__":
    sys.exit(main())
